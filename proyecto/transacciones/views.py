"""
Vistas para el sistema de transacciones de Global Exchange.

Este módulo contiene todas las vistas necesarias para el procesamiento de transacciones
de compra y venta de monedas extranjeras, incluyendo la gestión del flujo completo
desde la selección inicial hasta la confirmación final.

Funcionalidades principales:
    - Proceso completo de compra de monedas (4 pasos)
    - Proceso completo de venta de monedas (4 pasos)
    - Gestión de recargos por medio de pago
    - Historial y consulta de transacciones
    - Validaciones de límites en tiempo real
    - Generación y gestión de tokens de seguridad

Author: Equipo de desarrollo Global Exchange
Date: 2025
"""

import logging
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.conf import settings
from django.http import JsonResponse
from monedas.models import Moneda, StockGuaranies
from medios_acreditacion.models import TarjetaLocal
from .forms import SeleccionMonedaMontoForm, VariablesForm
from .models import Transaccion, Recargos, LimiteGlobal, Tauser, calcular_conversion, procesar_pago_stripe, procesar_transaccion, verificar_cambio_cotizacion_sesion, generar_token_transaccion, generar_factura_electronica, verificar_factura, descargar_factura
from .utils_2fa import is_2fa_enabled
from decimal import Decimal
from clientes.models import Cliente
import ast
import stripe
import logging
from datetime import date, timedelta
from django.utils import timezone
from django.db import models
import stripe

# Importaciones para exportación de archivos (se importan en las funciones para manejo de errores)
# from reportlab.pdfgen import canvas
# from reportlab.lib.pagesizes import letter, A4
# from reportlab.lib.styles import getSampleStyleSheet
# from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
# from reportlab.lib import colors
# from reportlab.lib.units import inch
# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from io import BytesIO

# Configurar Stripe
stripe.api_key = settings.STRIPE_SECRET_KEY
logger = logging.getLogger(__name__)

# ============================================================================
# PROCESO DE COMPRA DE MONEDAS
# ============================================================================

@login_required
def compra_monto_moneda(request):
    """
    Primer paso del proceso de compra: selección de moneda y monto.
    
    Permite al usuario seleccionar la moneda extranjera que desea comprar
    y especificar el monto. Incluye validaciones de límites de transacción
    antes de proceder al siguiente paso.
    
    Validaciones realizadas:
        - Usuario debe tener un cliente activo
        - Monto debe cumplir con límites diarios y mensuales
        - Moneda debe estar activa en el sistema
    
    Args:
        request (HttpRequest): Petición HTTP con datos del formulario
        
    Returns:
        HttpResponse: Renderiza formulario o redirecciona al siguiente paso
        
    Template:
        transacciones/seleccion_moneda_monto.html
        
    Context:
        - form: Formulario de selección de moneda y monto
        - paso_actual: Número del paso actual (1)
        - total_pasos: Total de pasos en el proceso (4)
        - titulo_paso: Título descriptivo del paso
        - tipo_transaccion: Tipo de operación ('compra')
        - limites_disponibles: Información de límites del cliente
    """
    transacciones_pasadas = Transaccion.objects.filter(usuario=request.user, estado='Pendiente')
    if transacciones_pasadas:
        for t in transacciones_pasadas:
            if t.fecha_hora < timezone.now() - timedelta(minutes=5):
                t.estado = 'Cancelada'
                t.razon = 'Expira el tiempo para confirmar la transacción'
                t.save()
    if request.session.get('transaccion_id'):
        t = Transaccion.objects.filter(id=request.session.get('transaccion_id')).first()
        t.estado = 'Cancelada'
        t.razon = 'Usuario sale del proceso de compra'
        t.save()
        del request.session['transaccion_id']
    if request.method == 'POST':
        form = SeleccionMonedaMontoForm(request.POST)
        if form.is_valid():
            moneda = form.cleaned_data['moneda']
            monto = form.cleaned_data['monto_decimal']
            cliente_activo = request.user.cliente_activo
            
            # Si pasa las validaciones, guardar los datos en la sesión
            request.session['compra_datos'] = {
                'moneda': moneda.id,
                'monto': str(monto),  # Convertir Decimal a string para serialización
                'paso_actual': 2
            }
            # Guardar precios iniciales en la sesión
            precios = moneda.get_precios_cliente(cliente_activo)
            request.session['precio_venta_inicial'] = precios['precio_venta']
            # Redireccionar al siguiente paso sin parámetros en la URL
            return redirect('transacciones:compra_medio_pago')
    else:
        # Verificar si el usuario tiene un cliente activo
        if not request.user.cliente_activo:
            messages.error(request, 'Debes tener un cliente activo para realizar compras.')
            return redirect('inicio')
        if request.user.cliente_activo.ultimo_consumo:
            if request.user.cliente_activo.ultimo_consumo != date.today():
                request.user.cliente_activo.consumo_diario = 0
                request.user.cliente_activo.save()
            if request.user.cliente_activo.ultimo_consumo.month != date.today().month:
                request.user.cliente_activo.consumo_mensual = 0
                request.user.cliente_activo.save()

        # Si cliente_activo.ultimo_consumo es None, no hacemos nada y los consumos
        # se mantienen en 0 (que es el valor inicial/correcto para un nuevo día/mes).
        form = SeleccionMonedaMontoForm()
    
    context = {
        'form': form,
        'disponible_diario': (LimiteGlobal.objects.first().limite_diario - request.user.cliente_activo.consumo_diario),
        'disponible_mensual': (LimiteGlobal.objects.first().limite_mensual - request.user.cliente_activo.consumo_mensual),
        'paso_actual': 1,
        'titulo': 'Selección de Moneda y Monto',
        'tipo_transaccion': 'compra'
    }
    
    return render(request, 'transacciones/seleccion_moneda_monto.html', context)

@login_required
def compra_medio_pago(request):
    """
    Segundo paso del proceso de compra: selección del medio de pago.
    
    Permite al usuario seleccionar cómo va a pagar por la moneda extranjera.
    Las opciones incluyen efectivo, billetera electrónica, transferencia
    bancaria y tarjetas de crédito registradas en Stripe.
    
    Validaciones realizadas:
        - Datos del paso anterior deben existir en sesión
        - Usuario debe tener cliente activo
        - Medio de pago debe estar disponible para el cliente
        - Verificación de cambios en cotización
    
    Args:
        request (HttpRequest): Petición HTTP con selección de medio de pago
        
    Returns:
        HttpResponse: Renderiza formulario o redirecciona al siguiente paso
        
    Template:
        transacciones/seleccion_medio_pago.html o transacciones/cotizacion_cambiada.html
        
    Context:
        - moneda: Moneda seleccionada en el paso anterior
        - monto: Monto seleccionado en el paso anterior
        - medios_pago: Lista de medios de pago disponibles
        - medio_pago_seleccionado: Medio actualmente seleccionado
        - cliente_activo: Cliente activo del usuario
        - paso_actual: Número del paso actual (2)
        - total_pasos: Total de pasos en el proceso (4)
        - titulo_paso: Título descriptivo del paso
        - tipo_transaccion: Tipo de operación ('compra')
    """
    # Verificar que el usuario tenga un cliente activo
    if not request.user.cliente_activo:
        messages.error(request, 'Debe tener un cliente activo seleccionado para continuar.')
        return redirect('inicio')
    
    # Verificar que existan datos del paso anterior
    compra_datos = request.session.get('compra_datos')
    if not compra_datos or compra_datos.get('paso_actual') != 2:
        messages.error(request, 'Debe completar el primer paso antes de continuar.')
        return redirect('transacciones:compra_monto_moneda')
    
    # Recuperar los datos de la sesión
    try:
        moneda = Moneda.objects.get(id=compra_datos['moneda'])
        monto = Decimal(compra_datos['monto'])
    except (Moneda.DoesNotExist, ValueError, KeyError):
        messages.error(request, 'Error al recuperar los datos. Reinicie el proceso.')
        return redirect('transacciones:compra_monto_moneda')
    
    if request.method == 'POST':
        # Verificar si es selección de medio de pago o avance al siguiente paso
        accion = request.POST.get('accion')
        
        if accion == 'seleccionar_medio':
            # Manejar la selección de medio de pago
            medio_pago = request.POST.get('medio_pago_id')
            if medio_pago:
                try:
                    # Actualizar los datos de la sesión (sin cambiar el paso_actual)
                    compra_datos.update({
                        'medio_pago': medio_pago
                    })
                    request.session['compra_datos'] = compra_datos
                    if medio_pago.startswith('{'):
                        medio_pago_dict = ast.literal_eval(medio_pago)
                        messages.success(request, f'Medio de pago Tarjeta de Crédito (**** **** **** {medio_pago_dict["last4"]}) seleccionado correctamente.')
                    else:
                        messages.success(request, f'Medio de pago {medio_pago} seleccionado correctamente.')
                    return redirect('transacciones:compra_medio_pago')  # Permanecer en el mismo paso
        
                except Exception as e:
                    messages.error(request, 'Error al seleccionar el medio de pago. Intente nuevamente.')
                    return redirect('transacciones:compra_medio_pago')
        
        elif accion == 'continuar':
            # Verificar que hay un medio de pago seleccionado
            if not compra_datos.get('medio_pago'):
                messages.error(request, 'Debe seleccionar un medio de pago antes de continuar.')
                return redirect('transacciones:compra_medio_pago')
            
            # Actualizar el paso actual y continuar al siguiente paso
            compra_datos['paso_actual'] = 3
            request.session['compra_datos'] = compra_datos
            return redirect('transacciones:compra_medio_cobro')
    
    medios_pago_disponibles = [
        'Efectivo',
        'Transferencia Bancaria'
    ]
    # Agregar billeteras electrónicas como medio de pago si hay recargos configurados
    recargos_billetera = Recargos.objects.filter(medio='Billetera Electrónica')
    for recargo in recargos_billetera:
        medios_pago_disponibles.append(recargo.marca)
    # Verificar si el cliente tiene tarjetas de crédito activas en Stripe
    if request.user.cliente_activo.tiene_tarjetas_activas():
        for tarjeta in request.user.cliente_activo.obtener_tarjetas_stripe():
            medios_pago_disponibles.append(tarjeta)
    if TarjetaLocal.objects.filter(cliente=request.user.cliente_activo, activo=True).exists():
        for tarjeta in TarjetaLocal.objects.filter(cliente=request.user.cliente_activo, activo=True):
            dict_tarjeta = {
                "id": tarjeta.id,
                "brand": tarjeta.brand,
                "last4": tarjeta.last4
            }
            medios_pago_disponibles.append(dict_tarjeta)
    # Obtener el medio de pago seleccionado actualmente (si hay uno)
    medio_pago_seleccionado = None
    if compra_datos.get('medio_pago'):
        if compra_datos['medio_pago'].startswith('{'):
            medio_pago_seleccionado = ast.literal_eval(compra_datos['medio_pago'])
        else:
            medio_pago_seleccionado = compra_datos['medio_pago']

    context = {
        'moneda': moneda,
        'monto': monto,
        'medios_pago': medios_pago_disponibles,
        'medio_pago_seleccionado': medio_pago_seleccionado,
        'cliente_activo': request.user.cliente_activo,
        'paso_actual': 2,
        'titulo_paso': 'Selección de Medio de Pago',
        'tipo_transaccion': 'compra'  # Agregar contexto para diferenciar en plantilla
    }
    
    return render(request, 'transacciones/seleccion_medio_pago.html', context)

@login_required
def compra_medio_cobro(request):
    """
    Tercer paso del proceso de compra: selección del medio de cobro.
    
    Permite al usuario seleccionar cómo va a recibir la moneda extranjera
    que está comprando. Actualmente solo se ofrece la opción de efectivo
    como medio de cobro para las compras.
    
    Validaciones realizadas:
        - Datos de pasos anteriores deben existir en sesión
        - Usuario debe tener cliente activo
        - Medio de cobro debe estar disponible
        - Verificación de cambios en cotización
    
    Args:
        request (HttpRequest): Petición HTTP con selección de medio de cobro
        
    Returns:
        HttpResponse: Renderiza formulario o redirecciona al siguiente paso
        
    Template:
        transacciones/seleccion_medio_cobro.html o transacciones/cotizacion_cambiada.html
        
    Context:
        - moneda: Moneda seleccionada
        - monto: Monto seleccionado
        - medio_pago: Medio de pago seleccionado
        - medios_cobro: Lista de medios de cobro disponibles
        - medio_cobro_seleccionado: Medio de cobro actualmente seleccionado
        - cliente_activo: Cliente activo del usuario
        - paso_actual: Número del paso actual (3)
        - total_pasos: Total de pasos en el proceso (4)
        - titulo_paso: Título descriptivo del paso
        - tipo_transaccion: Tipo de operación ('compra')
    """
    # Verificar que el usuario tenga un cliente activo
    if not request.user.cliente_activo:
        messages.error(request, 'Debe tener un cliente activo seleccionado para continuar.')
        return redirect('inicio')
    
    # Verificar que existan datos del paso anterior
    compra_datos = request.session.get('compra_datos')
    if not compra_datos or compra_datos.get('paso_actual') != 3:
        messages.error(request, 'Debe completar el segundo paso antes de continuar.')
        return redirect('transacciones:compra_medio_pago')
    
    # Recuperar los datos de la sesión
    try:
        moneda = Moneda.objects.get(id=compra_datos['moneda'])
        # Guardar valores iniciales de cotización
        request.session['tasa_base_inicial'] = moneda.tasa_base
        request.session['comision_compra_inicial'] = moneda.comision_compra
        request.session['comision_venta_inicial'] = moneda.comision_venta
        monto = Decimal(compra_datos['monto'])
        medio_pago = compra_datos['medio_pago']
    except (Moneda.DoesNotExist, ValueError, KeyError):
        messages.error(request, 'Error al recuperar los datos. Reinicie el proceso.')
        return redirect('transacciones:compra_monto_moneda')
    
    if request.method == 'POST':
        # Verificar si es selección de medio de cobro o avance al siguiente paso
        accion = request.POST.get('accion')
        
        if accion == 'seleccionar_medio':
            # Manejar la selección de medio de cobro
            medio_cobro = request.POST.get('medio_cobro_id')
            if medio_cobro:
                try:
                    # Actualizar los datos de la sesión (sin cambiar el paso_actual)
                    compra_datos.update({
                        'medio_cobro': medio_cobro
                    })
                    request.session['compra_datos'] = compra_datos

                    messages.success(request, f'Medio de cobro {medio_cobro} seleccionado correctamente.')
                    return redirect('transacciones:compra_medio_cobro')  # Permanecer en el mismo paso
        
                except Exception as e:
                    messages.error(request, 'Error al seleccionar el medio de cobro. Intente nuevamente.')
                    return redirect('transacciones:compra_medio_cobro')
        
        elif accion == 'continuar':
            # Verificar que hay un medio de cobro seleccionado
            if not compra_datos.get('medio_cobro'):
                messages.error(request, 'Debe seleccionar un medio de cobro antes de continuar.')
                return redirect('transacciones:compra_medio_cobro')
            
            # Actualizar el paso actual y continuar al siguiente paso
            compra_datos['paso_actual'] = 4
            request.session['compra_datos'] = compra_datos
            return redirect('transacciones:compra_confirmacion')
    
    # Construir lista de medios de cobro disponibles
    medios_cobro_disponibles = ['Efectivo']  # Opción fija

    # Obtener el medio de cobro seleccionado actualmente (si hay uno)
    medio_cobro_seleccionado = None
    if compra_datos.get('medio_cobro'):
        medio_cobro_seleccionado = compra_datos['medio_cobro']

    context = {
        'moneda': moneda,
        'monto': monto,
        'medio_pago': medio_pago,
        'medios_cobro': medios_cobro_disponibles,
        'medio_cobro_seleccionado': medio_cobro_seleccionado,
        'cliente_activo': request.user.cliente_activo,
        'paso_actual': 3,
        'titulo_paso': 'Selección de Medio de Cobro',
        'tipo_transaccion': 'compra'
    }
    
    return render(request, 'transacciones/seleccion_medio_cobro.html', context)

@login_required
def compra_confirmacion(request):
    """
    Cuarto paso del proceso de compra: confirmación y creación de transacción.
    
    Muestra un resumen completo de la transacción y procede a crearla en
    la base de datos. Para medios de pago como Efectivo o Transferencia,
    genera un token de seguridad con validez de 5 minutos.
    
    Acciones realizadas:
        - Verificación de cambios en cotización
        - Creación de registro de transacción en base de datos
        - Generación de token para medios específicos
        - Configuración del estado inicial como 'Pendiente'
        - Vinculación con cliente y usuario activos
    
    Args:
        request (HttpRequest): Petición HTTP de confirmación
        
    Returns:
        HttpResponse: Renderiza página de confirmación o modal de cambio
        
    Template:
        transacciones/confirmacion.html o transacciones/cotizacion_cambiada.html
        
    Context:
        - moneda: Moneda de la transacción
        - monto: Monto de la transacción
        - medio_pago: Medio de pago seleccionado
        - medio_cobro: Medio de cobro seleccionado
        - cliente_activo: Cliente que realiza la transacción
        - transaccion: Instancia de transacción creada
        - paso_actual: Número del paso actual (4)
        - total_pasos: Total de pasos en el proceso (4)
        - titulo_paso: Título descriptivo del paso
        - tipo_transaccion: Tipo de operación ('compra')
    """
    # Verificar que el usuario tenga un cliente activo
    if not request.user.cliente_activo:
        messages.error(request, 'Debe tener un cliente activo seleccionado para continuar.')
        return redirect('inicio')
    
    # Verificar que existan datos del paso anterior
    compra_datos = request.session.get('compra_datos')
    if not compra_datos or compra_datos.get('paso_actual') != 4:
        messages.error(request, 'Debe completar el tercer paso antes de continuar.')
        return redirect('transacciones:compra_medio_cobro')

    if request.method == 'POST':
        accion = request.POST.get('accion')
        action = request.POST.get('action')
        if accion == 'confirmar':
            cambios = verificar_cambio_cotizacion_sesion(request, 'compra')
            if cambios and cambios.get('hay_cambios'):
                transaccion = Transaccion.objects.filter(id=request.session.get('transaccion_id')).first()
                datos_transaccion = calcular_conversion(transaccion.monto, transaccion.moneda, 'compra', transaccion.medio_pago, transaccion.medio_cobro, request.user.cliente_activo.segmento)
                transaccion.precio_base = datos_transaccion['precio_base']
                transaccion.cotizacion = datos_transaccion['cotizacion']
                transaccion.beneficio_segmento = datos_transaccion['beneficio_segmento']
                transaccion.porc_beneficio_segmento = datos_transaccion['porc_beneficio_segmento']
                transaccion.recargo_pago = datos_transaccion['monto_recargo_pago']
                transaccion.porc_recargo_pago = datos_transaccion['porc_recargo_pago']
                transaccion.recargo_cobro = datos_transaccion['monto_recargo_cobro']
                transaccion.porc_recargo_cobro = datos_transaccion['porc_recargo_cobro']
                transaccion.redondeo_efectivo_monto = datos_transaccion['redondeo_efectivo_monto']
                transaccion.redondeo_efectivo_precio_final = datos_transaccion['redondeo_efectivo_precio_final']
                transaccion.monto_original = datos_transaccion['monto_original']
                transaccion.monto = datos_transaccion['monto']
                transaccion.precio_final = datos_transaccion['precio_final']
                transaccion.save()
                context = {
                    'cambios': cambios,
                    'transaccion': transaccion,
                    'paso_actual': 4,
                    'titulo_paso': 'Confirmación de Compra',
                    'enable_2fa': is_2fa_enabled(),
                    'user_email': request.user.email,
                    'has_email': bool(request.user.email)
                }
                
                return render(request, 'transacciones/confirmacion.html', context)
            return redirect('transacciones:compra_exito')
        elif accion == 'cancelar':
            messages.info(request, 'Has cancelado la transacción.')
            t = Transaccion.objects.filter(id=request.session.get('transaccion_id')).first()
            if t:
                t.estado = 'Cancelada'
                t.razon = 'Usuario canceló la transacción en el formulario'
                t.save()
                # Limpiar datos de sesión
            if 'compra_datos' in request.session:
                del request.session['compra_datos']
            if 'precio_venta_inicial' in request.session:
                del request.session['precio_venta_inicial']
            if 'transaccion_id' in request.session:
                del request.session['transaccion_id']
            return redirect('inicio')
        elif action == 'aceptar':
            transaccion = Transaccion.objects.get(id=request.session.get('transaccion_id'))
            precios_actuales = transaccion.moneda.get_precios_cliente(request.user.cliente_activo)
            request.session['precio_venta_inicial'] = precios_actuales['precio_venta']
            messages.success(request, 'Precios actualizados. Continuando con la transacción.')
            context = {
                'transaccion': transaccion,
                'paso_actual': 4,
                'titulo_paso': 'Confirmación de Compra',
                'enable_2fa': is_2fa_enabled(),
                'user_email': request.user.email,
                'has_email': bool(request.user.email)
            }
            
            return render(request, 'transacciones/confirmacion.html', context)
        elif action == 'cancelar':
            t = Transaccion.objects.filter(id=request.session.get('transaccion_id')).first()
            if t:
                t.estado = 'Cancelada'
                t.razon = 'Usuario cancela debido a cambios de cotización'
                t.save()
                # Limpiar datos de sesión
            if 'compra_datos' in request.session:
                del request.session['compra_datos']
            if 'precio_venta_inicial' in request.session:
                del request.session['precio_venta_inicial']
            if 'transaccion_id' in request.session:
                del request.session['transaccion_id']
            messages.info(request, 'Transacción cancelada debido a cambios en la cotización.')
            return redirect('inicio')
    else:   
        # Recuperar los datos de la sesión
        try:
            moneda = Moneda.objects.get(id=compra_datos['moneda'])
            monto = Decimal(compra_datos['monto'])
            medio_pago = compra_datos['medio_pago']
            medio_cobro = compra_datos['medio_cobro']
        except (Moneda.DoesNotExist, ValueError, KeyError):
            messages.error(request, 'Error al recuperar los datos. Reinicie el proceso.')
            return redirect('transacciones:compra_monto_moneda')
        if request.session.get('transaccion_id'):
            transaccion = Transaccion.objects.filter(id=request.session.get('transaccion_id')).first()
            context = {
            'transaccion': transaccion,
            'paso_actual': 4,
            'titulo_paso': 'Confirmación de Compra'
            }
            return render(request, 'transacciones/confirmacion.html', context)
        
        # Crear la transacción en la base de datos
        try:
            if medio_pago.startswith('{'):
                medio_pago_dict = ast.literal_eval(medio_pago)
                str_medio_pago = f'Tarjeta de Crédito (**** **** **** {medio_pago_dict["last4"]})'
            else:
                str_medio_pago = medio_pago
            datos_transaccion = calcular_conversion(monto, moneda, 'compra', medio_pago, medio_cobro, request.user.cliente_activo.segmento)
            if datos_transaccion['precio_final'] > (LimiteGlobal.objects.first().limite_diario - request.user.cliente_activo.consumo_diario) or datos_transaccion['precio_final'] > (LimiteGlobal.objects.first().limite_mensual - request.user.cliente_activo.consumo_mensual):
                messages.warning(request, 'El monto final excede sus límites diarios o mensuales. Reinicie el proceso con un monto menor.')
                return redirect('transacciones:compra_monto_moneda')
            transaccion = Transaccion.objects.create(
                cliente=request.user.cliente_activo,
                tipo='compra',
                moneda=moneda,
                monto=datos_transaccion['monto'],
                monto_original=datos_transaccion['monto_original'],
                cotizacion=datos_transaccion['cotizacion'],
                precio_base=datos_transaccion['precio_base'],
                beneficio_segmento=datos_transaccion['beneficio_segmento'],
                porc_beneficio_segmento=datos_transaccion['porc_beneficio_segmento'],
                recargo_pago=datos_transaccion['monto_recargo_pago'],
                porc_recargo_pago=datos_transaccion['porc_recargo_pago'],
                recargo_cobro=datos_transaccion['monto_recargo_cobro'],
                porc_recargo_cobro=datos_transaccion['porc_recargo_cobro'],
                redondeo_efectivo_monto=datos_transaccion['redondeo_efectivo_monto'],
                redondeo_efectivo_precio_final=datos_transaccion['redondeo_efectivo_precio_final'],
                precio_final=datos_transaccion['precio_final'],
                medio_pago=str_medio_pago,
                medio_cobro=medio_cobro,
                usuario=request.user
            )
            request.session['transaccion_id'] = transaccion.id
                
        except Exception as e:
            messages.error(request, 'Error al crear la transacción. Intente nuevamente.')
            print(e)
            return redirect('transacciones:compra_medio_cobro')
        
        cambios = verificar_cambio_cotizacion_sesion(request, 'compra')
        if cambios and cambios.get('hay_cambios'):
            context = {
                'cambios': cambios,
                'transaccion': transaccion,
                'paso_actual': 4,
                'titulo_paso': 'Confirmación de Compra',
                'enable_2fa': is_2fa_enabled(),
                'user_email': request.user.email,
                'has_email': bool(request.user.email)
            }
            
            return render(request, 'transacciones/confirmacion.html', context)
        
        context = {
            'transaccion': transaccion,
            'paso_actual': 4,
            'titulo_paso': 'Confirmación de Compra',
            'enable_2fa': is_2fa_enabled(),
            'user_email': request.user.email,
            'has_email': bool(request.user.email)
        }
        
        return render(request, 'transacciones/confirmacion.html', context)


@login_required
def compra_exito(request, token=None):
    """
    Página final del proceso de compra: mensaje de éxito.
    
    Verifica si hay cambios en la cotización antes de finalizar la transacción.
    Si hay cambios, muestra el modal de confirmación. Si no hay cambios o el usuario
    acepta los nuevos precios, muestra confirmación de éxito.
    
    Args:
        request (HttpRequest): Petición HTTP
        token (str, optional): Token de la transacción para retorno de la pasarela
        
    Returns:
        HttpResponse: Página de éxito o modal de cambio de cotización
        
    Template:
        transacciones/exito.html o transacciones/cotizacion_cambiada.html
    """
    # Verificar que el usuario tenga un cliente activo
    if not request.user.cliente_activo:
        messages.error(request, 'Debe tener un cliente activo seleccionado para continuar.')
        return redirect('inicio')

    transaccion = None
    # Si viene token desde la pasarela, buscar la transacción por token
    if token:
        try:
            transaccion = Transaccion.objects.get(token=token)
            
            # Si viene desde la pasarela, validar los datos del pago
            if transaccion.estado == 'Pendiente' and request.GET.get('estado') == 'exito':
                # Datos enviados por la pasarela
                try:
                    monto_pago = float(request.GET.get('monto', 0))
                    cuenta_pago = request.GET.get('nro_cuenta')
                    banco_pago = request.GET.get('banco')
                    
                    # Verificación de cuenta/billetera
                    CUENTAS_GLOBAL_EXCHANGE = {
                        'Transferencia Bancaria': 'SUDAMERIS-123456',
                        'Tigo Money': 'TIGO-0981000001-1010101',
                        'Billetera Personal': 'PERSONAL-0982000002-2020202',
                        'Zimple': 'ZIMPLE-0983000003-3030303'
                    }
                    # Verificación del monto
                    monto_valido = monto_pago >= float(transaccion.precio_final)
                    
                    if transaccion.medio_pago == "Transferencia Bancaria":
                        # Para transferencias, el CUENTAS_GLOBAL_EXCHANGE ya tiene banco+cuenta
                        cuenta_valida = f"{banco_pago}-{cuenta_pago}" == CUENTAS_GLOBAL_EXCHANGE.get(transaccion.medio_pago)
                    else:
                        # Para billeteras, la cuenta completa ya incluye el prefijo del banco
                        cuenta_valida = cuenta_pago == CUENTAS_GLOBAL_EXCHANGE.get(transaccion.medio_pago)

                
                    if monto_valido and cuenta_valida:
                        transaccion.estado = 'Confirmada'
                        generar_factura_electronica(transaccion)
                        transaccion.fecha_hora = timezone.now()
                        transaccion.pagado = transaccion.precio_final
                        transaccion.save()
                        transaccion.cliente.consumo_diario += transaccion.precio_final
                        transaccion.cliente.consumo_mensual += transaccion.precio_final
                        transaccion.cliente.ultimo_consumo = date.today()
                        transaccion.cliente.save()
                        guaranies = StockGuaranies.objects.first()
                        guaranies.cantidad += transaccion.precio_final - transaccion.recargo_pago
                        guaranies.save()
                        messages.success(request, 'Pago confirmado exitosamente')
                        request.session.pop('transaccion_id', None)
                        context = {
                            'transaccion': transaccion
                        }
                        return render(request, 'transacciones/exito.html', context)
                    else:
                        messages.error(request, 'Error en la validación del pago, se revierte el pago. Por favor, verifique los datos y el monto, e intente nuevamente.')
                        context = {
                            'transaccion': transaccion
                        }
                        return render(request, 'transacciones/informacion_pago.html', context)
                        
                except (ValueError, TypeError, AttributeError) as e:
                    transaccion.estado = 'Cancelada'
                    transaccion.razon = f'Error procesando datos del pago: {str(e)}'
                    transaccion.save()
                    messages.error(request, 'Error al procesar los datos del pago')
                    return redirect('inicio')
                    
            elif request.GET.get('estado') == 'error':
                transaccion.estado = 'Cancelada'
                transaccion.razon = 'Error reportado por la pasarela de pago'
                transaccion.save()
                messages.error(request, 'Error al procesar el pago')
                
        except Transaccion.DoesNotExist:
            messages.error(request, 'Token de transacción inválido.')
            return redirect('inicio')
    else:
        # Flujo normal: buscar por sesión
        try:
            idt = request.session.get('transaccion_id')
            transaccion = Transaccion.objects.get(id=idt)
        except:
            return redirect('inicio')

    compra_datos = request.session.get('compra_datos')
    if not compra_datos:
        messages.error(request, 'Debe completar el cuarto paso antes de continuar.')
        return redirect('transacciones:compra_medio_cobro')
    try:
        medio_pago = compra_datos['medio_pago']
    except (Moneda.DoesNotExist, ValueError, KeyError):
        messages.error(request, 'Error al recuperar los datos. Reinicie el proceso.')
        return redirect('transacciones:compra_monto_moneda')

    if transaccion.token is None:
        if medio_pago.startswith('{'):
            medio_pago_dict = ast.literal_eval(medio_pago)
            if medio_pago_dict['brand'] in ['VISA','MASTERCARD']:
                pago = procesar_pago_stripe(transaccion, medio_pago_dict["id"])
                if pago['success']:
                    messages.success(request, 'Pago con tarjeta de crédito procesado exitosamente.')
                    procesar_transaccion(transaccion)
                    generar_token_transaccion(transaccion)
                else:
                    transaccion.estado = 'Cancelada'
                    transaccion.razon = 'Error en el procesamiento del pago con tarjeta de crédito'
                    transaccion.save()
                    messages.error(request, 'Error al procesar el pago con tarjeta de crédito. Intente nuevamente.')
                    return redirect('transacciones:compra_monto_moneda')
            else:
                messages.success(request, 'Pago con tarjeta de crédito procesado exitosamente.')
                procesar_transaccion(transaccion)
                generar_token_transaccion(transaccion)
        else:
            try:
                generar_token_transaccion(transaccion)
            except Exception as e:
                messages.error(request, 'Error al generar token de transacción. Intente nuevamente.')
                return redirect('transacciones:compra_medio_cobro')
            transaccion.save()
            if transaccion.medio_pago != 'Efectivo':
                context = {
                    'transaccion': transaccion
                }
                return render(request, 'transacciones/informacion_pago.html', context)

    else:
        return redirect('inicio')

    # Limpiar sesión
    request.session.pop('transaccion_id', None)

    context = {
        'transaccion': transaccion
    }

    return render(request, 'transacciones/exito.html', context)


# ============================================================================
# PROCESO DE VENTA DE MONEDAS
# ============================================================================

@login_required
def venta_monto_moneda(request):
    """
    Primer paso del proceso de venta: selección de moneda y monto.
    
    Permite al usuario seleccionar la moneda extranjera que desea vender
    y especificar el monto. Similar al proceso de compra pero con validaciones
    específicas para operaciones de venta.
    
    Validaciones realizadas:
        - Usuario debe tener un cliente activo
        - Monto debe cumplir con límites diarios y mensuales
        - Moneda debe estar activa en el sistema
    
    Args:
        request (HttpRequest): Petición HTTP con datos del formulario
        
    Returns:
        HttpResponse: Renderiza formulario o redirecciona al siguiente paso
        
    Template:
        transacciones/seleccion_moneda_monto.html
        
    Context:
        - form: Formulario de selección de moneda y monto
        - paso_actual: Número del paso actual (1)
        - total_pasos: Total de pasos en el proceso (4)
        - titulo_paso: Título descriptivo del paso
        - tipo_transaccion: Tipo de operación ('venta')
        - limites_disponibles: Información de límites del cliente
    """
    transacciones_pasadas = Transaccion.objects.filter(usuario=request.user, estado='Pendiente')
    if transacciones_pasadas:
        for t in transacciones_pasadas:
            if t.fecha_hora < timezone.now() - timedelta(minutes=5):
                t.estado = 'Cancelada'
                t.razon = 'Expira el tiempo para confirmar la transacción'
                t.save()
    if request.session.get('transaccion_id'):
        t = Transaccion.objects.filter(id=request.session.get('transaccion_id')).first()
        t.estado = 'Cancelada'
        t.razon = 'Usuario sale del proceso de compra'
        t.save()
        del request.session['transaccion_id']
    if request.method == 'POST':
        form = SeleccionMonedaMontoForm(request.POST)
        if form.is_valid():
            moneda = form.cleaned_data['moneda']
            monto = form.cleaned_data['monto_decimal']
            cliente_activo = request.user.cliente_activo
            # Si pasa las validaciones, guardar los datos en la sesión
            request.session['venta_datos'] = {
                'moneda': moneda.id,
                'monto': str(monto),  # Convertir Decimal a string para serialización
                'paso_actual': 2
            }
            # Guardar precios iniciales en la sesión
            precios = moneda.get_precios_cliente(cliente_activo)
            request.session['precio_compra_inicial'] = precios['precio_compra']
            # Redireccionar al siguiente paso sin parámetros en la URL
            return redirect('transacciones:venta_medio_pago')
    else:
        # Verificar si el usuario tiene un cliente activo
        if not request.user.cliente_activo:
            messages.error(request, 'Debes tener un cliente activo para realizar compras.')
            return redirect('inicio')
         
        # 1. Verificar si el cliente tiene un último consumo registrado
        if request.user.cliente_activo.ultimo_consumo:
            if request.user.cliente_activo.ultimo_consumo != date.today():
                request.user.cliente_activo.consumo_diario = 0
                request.user.cliente_activo.save()
            if request.user.cliente_activo.ultimo_consumo.month != date.today().month:
                request.user.cliente_activo.consumo_mensual = 0
                request.user.cliente_activo.save()
        # Si cliente_activo.ultimo_consumo es None, no hacemos nada y los consumos
        # se mantienen en 0 (que es el valor inicial/correcto para un nuevo día/mes).
        form = SeleccionMonedaMontoForm()
    
    context = {
        'form': form,
        'disponible_diario': (LimiteGlobal.objects.first().limite_diario - request.user.cliente_activo.consumo_diario),
        'disponible_mensual': (LimiteGlobal.objects.first().limite_mensual - request.user.cliente_activo.consumo_mensual),
        'paso_actual': 1,
        'titulo': 'Selección de Moneda y Monto',
        'tipo_transaccion': 'venta'
    }

    return render(request, 'transacciones/seleccion_moneda_monto.html', context)

@login_required
def venta_medio_pago(request):
    """
    Segundo paso del proceso de venta: selección del medio de pago.
    
    Permite al usuario seleccionar cómo va a recibir el pago por la moneda
    extranjera que está vendiendo. Para ventas, principalmente se ofrece
    efectivo, y para USD también tarjetas de crédito registradas.
    
    Validaciones realizadas:
        - Datos del paso anterior deben existir en sesión
        - Usuario debe tener cliente activo
        - Para tarjetas: solo disponibles para USD y clientes con tarjetas activas
    
    Args:
        request (HttpRequest): Petición HTTP con selección de medio de pago
        
    Returns:
        HttpResponse: Renderiza formulario o redirecciona al siguiente paso
        
    Template:
        transacciones/seleccion_medio_pago.html
        
    Context:
        - moneda: Moneda seleccionada en el paso anterior
        - monto: Monto seleccionado en el paso anterior
        - medios_pago: Lista de medios de pago disponibles
        - medio_pago_seleccionado: Medio actualmente seleccionado
        - cliente_activo: Cliente activo del usuario
        - paso_actual: Número del paso actual (2)
        - tipo_transaccion: Tipo de operación ('venta')
    """
    # Verificar que el usuario tenga un cliente activo
    if not request.user.cliente_activo:
        messages.error(request, 'Debe tener un cliente activo seleccionado para continuar.')
        return redirect('inicio')
    
    # Verificar que existan datos del paso anterior
    venta_datos = request.session.get('venta_datos')
    if not venta_datos or venta_datos.get('paso_actual') != 2:
        messages.error(request, 'Debe completar el primer paso antes de continuar.')
        return redirect('transacciones:venta_monto_moneda')
    
    # Recuperar los datos de la sesión
    try:
        moneda = Moneda.objects.get(id=venta_datos['moneda'])
        monto = Decimal(venta_datos['monto'])
    except (Moneda.DoesNotExist, ValueError, KeyError):
        messages.error(request, 'Error al recuperar los datos. Reinicie el proceso.')
        return redirect('transacciones:venta_monto_moneda')
    
    if request.method == 'POST':
        # Verificar si es selección de medio de pago o avance al siguiente paso
        accion = request.POST.get('accion')
        
        if accion == 'seleccionar_medio':
            # Manejar la selección de medio de pago
            medio_pago = request.POST.get('medio_pago_id')
            if medio_pago:
                try:
                    # Actualizar los datos de la sesión (sin cambiar el paso_actual)
                    venta_datos.update({
                        'medio_pago': medio_pago
                    })
                    request.session['venta_datos'] = venta_datos
                    if medio_pago.startswith('{'):
                        medio_pago_dict = ast.literal_eval(medio_pago)
                        messages.success(request, f'Medio de pago Tarjeta de Crédito (**** **** **** {medio_pago_dict["last4"]}) seleccionado correctamente.')
                    else:
                        messages.success(request, f'Medio de pago {medio_pago} seleccionado correctamente.')
                    return redirect('transacciones:venta_medio_pago')  # Permanecer en el mismo paso
        
                except Exception as e:
                    messages.error(request, 'Error al seleccionar el medio de pago. Intente nuevamente.')
                    return redirect('transacciones:venta_medio_pago')
        
        elif accion == 'continuar':
            # Verificar que hay un medio de pago seleccionado
            if not venta_datos.get('medio_pago'):
                messages.error(request, 'Debe seleccionar un medio de pago antes de continuar.')
                return redirect('transacciones:venta_medio_pago')
            
            # Actualizar el paso actual y continuar al siguiente paso
            venta_datos['paso_actual'] = 3
            request.session['venta_datos'] = venta_datos
            return redirect('transacciones:venta_medio_cobro')
    
    medios_pago_disponibles = [
        'Efectivo',
    ]
    # Para ventas, verificar tarjetas activas solo para USD
    if moneda.simbolo == 'USD' and request.user.cliente_activo.tiene_tarjetas_activas():
        for tarjeta in request.user.cliente_activo.obtener_tarjetas_stripe():
            medios_pago_disponibles.append(tarjeta)

    # Obtener el medio de pago seleccionado actualmente (si hay uno)
    medio_pago_seleccionado = None
    if venta_datos.get('medio_pago'):
        if venta_datos['medio_pago'].startswith('{'):
            medio_pago_seleccionado = ast.literal_eval(venta_datos['medio_pago'])
        else:
            medio_pago_seleccionado = venta_datos['medio_pago']

    context = {
        'moneda': moneda,
        'monto': monto,
        'medios_pago': medios_pago_disponibles,
        'medio_pago_seleccionado': medio_pago_seleccionado,
        'cliente_activo': request.user.cliente_activo,
        'paso_actual': 2,
        'titulo_paso': 'Selección de Medio de Pago',
        'tipo_transaccion': 'venta'  # Agregar contexto para diferenciar en plantilla
    }
    
    return render(request, 'transacciones/seleccion_medio_pago.html', context)

@login_required
def venta_medio_cobro(request):
    """
    Tercer paso del proceso de venta: selección del medio de cobro.
    
    Permite al usuario seleccionar cómo va a entregar la moneda extranjera
    que está vendiendo. Incluye opciones como efectivo, cuentas bancarias
    registradas y billeteras electrónicas del cliente.
    
    Validaciones realizadas:
        - Datos de pasos anteriores deben existir en sesión
        - Usuario debe tener cliente activo
        - Medios disponibles según configuración del cliente
    
    Args:
        request (HttpRequest): Petición HTTP con selección de medio de cobro
        
    Returns:
        HttpResponse: Renderiza formulario o redirecciona al siguiente paso
        
    Template:
        transacciones/seleccion_medio_cobro.html
        
    Context:
        - moneda: Moneda seleccionada
        - monto: Monto seleccionado
        - medio_pago: Medio de pago seleccionado
        - medios_cobro: Lista de medios de cobro disponibles (efectivo, cuentas, billeteras)
        - medio_cobro_seleccionado: Medio de cobro actualmente seleccionado
        - cliente_activo: Cliente activo del usuario
        - paso_actual: Número del paso actual (3)
        - tipo_transaccion: Tipo de operación ('venta')
    """
    # Verificar que el usuario tenga un cliente activo
    if not request.user.cliente_activo:
        messages.error(request, 'Debe tener un cliente activo seleccionado para continuar.')
        return redirect('inicio')
    # Verificar que existan datos del paso anterior
    venta_datos = request.session.get('venta_datos')
    if not venta_datos or venta_datos.get('paso_actual') != 3:
        messages.error(request, 'Debe completar el segundo paso antes de continuar.')
        return redirect('transacciones:venta_medio_pago')
    
    # Recuperar los datos de la sesión
    try:
        moneda = Moneda.objects.get(id=venta_datos['moneda'])
        # Guardar valores iniciales de cotización
        request.session['tasa_base_inicial'] = moneda.tasa_base
        request.session['comision_compra_inicial'] = moneda.comision_compra
        request.session['comision_venta_inicial'] = moneda.comision_venta
        monto = Decimal(venta_datos['monto'])
        medio_pago = venta_datos['medio_pago']
    except (Moneda.DoesNotExist, ValueError, KeyError):
        messages.error(request, 'Error al recuperar los datos. Reinicie el proceso.')
        return redirect('transacciones:venta_monto_moneda')
    
    if request.method == 'POST':
        # Verificar si es selección de medio de cobro o avance al siguiente paso
        accion = request.POST.get('accion')
        
        if accion == 'seleccionar_medio':
            # Manejar la selección de medio de cobro
            medio_cobro = request.POST.get('medio_cobro_id')
            if medio_cobro:
                try:
                    # Actualizar los datos de la sesión (sin cambiar el paso_actual)
                    venta_datos.update({
                        'medio_cobro': medio_cobro
                    })
                    request.session['venta_datos'] = venta_datos
                    if medio_cobro.startswith('{'):
                        medio_cobro_dict = ast.literal_eval(medio_cobro)
                        messages.success(request, f'Medio de cobro {medio_cobro_dict["tipo_billetera"]} seleccionado correctamente.')
                    else:
                        messages.success(request, f'Medio de cobro {medio_cobro} seleccionado correctamente.')
                    return redirect('transacciones:venta_medio_cobro')  # Permanecer en el mismo paso
        
                except Exception as e:
                    messages.error(request, 'Error al seleccionar el medio de cobro. Intente nuevamente.')
                    return redirect('transacciones:venta_medio_cobro')
        
        elif accion == 'continuar':
            # Verificar que hay un medio de cobro seleccionado
            if not venta_datos.get('medio_cobro'):
                messages.error(request, 'Debe seleccionar un medio de cobro antes de continuar.')
                return redirect('transacciones:venta_medio_cobro')
            
            # Actualizar el paso actual y continuar al siguiente paso
            venta_datos['paso_actual'] = 4
            request.session['venta_datos'] = venta_datos
            return redirect('transacciones:venta_confirmacion')
    
    # Construir lista de medios de cobro disponibles
    medios_cobro_disponibles = ['Efectivo']  # Opción fija
    
    # Agregar cuentas bancarias si las hay
    cuentas_bancarias = request.user.cliente_activo.cuentas_bancarias.all()
    for cuenta in cuentas_bancarias:
        medio_descripcion = f"Cuenta bancaria - {cuenta.get_banco_display()} ({cuenta.numero_cuenta})"
        medios_cobro_disponibles.append(medio_descripcion)
    
    # Agregar billeteras si las hay
    billeteras = request.user.cliente_activo.billeteras.all()
    for billetera in billeteras:
        billetera_dict = {
            'id': billetera.id,
            'tipo_billetera': billetera.get_tipo_billetera_display(),
            'telefono': billetera.telefono,
            'nombre_titular': billetera.nombre_titular,
            'nro_documento': billetera.nro_documento
        }
        medios_cobro_disponibles.append(billetera_dict)

    # Obtener el medio de cobro seleccionado actualmente (si hay uno)
    medio_cobro_seleccionado = None
    if venta_datos.get('medio_cobro'):
        if venta_datos['medio_cobro'].startswith('{'):
            medio_cobro_seleccionado = ast.literal_eval(venta_datos['medio_cobro'])
        else:
            medio_cobro_seleccionado = venta_datos['medio_cobro']

    context = {
        'moneda': moneda,
        'monto': monto,
        'medio_pago': medio_pago,
        'medios_cobro': medios_cobro_disponibles,
        'medio_cobro_seleccionado': medio_cobro_seleccionado,
        'cliente_activo': request.user.cliente_activo,
        'paso_actual': 3,
        'titulo_paso': 'Selección de Medio de Cobro',
        'tipo_transaccion': 'venta'
    }
    
    return render(request, 'transacciones/seleccion_medio_cobro.html', context)

@login_required
def venta_confirmacion(request):
    """
    Cuarto paso del proceso de venta: confirmación y creación de transacción.
    
    Muestra un resumen completo de la transacción de venta y procede a crearla
    en la base de datos. Para ventas en efectivo, genera un token de seguridad
    con validez de 5 minutos.
    
    Acciones realizadas:
        - Creación de registro de transacción en base de datos
        - Generación de token para pagos en efectivo
        - Configuración del estado inicial como 'Pendiente'
        - Vinculación con cliente y usuario activos
    
    Args:
        request (HttpRequest): Petición HTTP de confirmación
        
    Returns:
        HttpResponse: Renderiza página de confirmación
        
    Template:
        transacciones/confirmacion.html
        
    Context:
        - moneda: Moneda de la transacción
        - monto: Monto de la transacción
        - medio_pago: Medio de pago seleccionado
        - medio_cobro: Medio de cobro seleccionado
        - cliente_activo: Cliente que realiza la transacción
        - transaccion: Instancia de transacción creada
        - paso_actual: Número del paso actual (4)
        - tipo_transaccion: Tipo de operación ('venta')
    """
    # Verificar que el usuario tenga un cliente activo
    if not request.user.cliente_activo:
        messages.error(request, 'Debe tener un cliente activo seleccionado para continuar.')
        return redirect('inicio')
    
    # Verificar que existan datos del paso anterior
    venta_datos = request.session.get('venta_datos')
    if not venta_datos or venta_datos.get('paso_actual') != 4:
        messages.error(request, 'Debe completar el tercer paso antes de continuar.')
        return redirect('transacciones:venta_medio_cobro')

    if request.method == 'POST':
        accion = request.POST.get('accion')
        action = request.POST.get('action')
        if accion == 'confirmar':
            cambios = verificar_cambio_cotizacion_sesion(request, 'venta')
            if cambios and cambios.get('hay_cambios'):
                transaccion = Transaccion.objects.filter(id=request.session.get('transaccion_id')).first()
                datos_transaccion = calcular_conversion(transaccion.monto, transaccion.moneda, 'venta', transaccion.medio_pago, transaccion.medio_cobro, request.user.cliente_activo.segmento)
                transaccion.precio_base = datos_transaccion['precio_base']
                transaccion.cotizacion = datos_transaccion['cotizacion']
                transaccion.beneficio_segmento = datos_transaccion['beneficio_segmento']
                transaccion.porc_beneficio_segmento = datos_transaccion['porc_beneficio_segmento']
                transaccion.recargo_pago = datos_transaccion['monto_recargo_pago']
                transaccion.porc_recargo_pago = datos_transaccion['porc_recargo_pago']
                transaccion.recargo_cobro = datos_transaccion['monto_recargo_cobro']
                transaccion.porc_recargo_cobro = datos_transaccion['porc_recargo_cobro']
                transaccion.redondeo_efectivo_monto = datos_transaccion['redondeo_efectivo_monto']
                transaccion.redondeo_efectivo_precio_final = datos_transaccion['redondeo_efectivo_precio_final']
                transaccion.monto_original = datos_transaccion['monto_original']
                transaccion.monto = datos_transaccion['monto']
                transaccion.precio_final = datos_transaccion['precio_final']
                transaccion.save()
                context = {
                    'cambios': cambios,
                    'transaccion': Transaccion.objects.get(id=request.session.get('transaccion_id')),
                    'paso_actual': 4,
                    'titulo_paso': 'Confirmación de Venta',
                    'enable_2fa': is_2fa_enabled(),
                    'user_email': request.user.email,
                    'has_email': bool(request.user.email)
                }
                
                return render(request, 'transacciones/confirmacion.html', context)
            return redirect('transacciones:venta_exito')
        elif accion == 'cancelar':
            messages.info(request, 'Has cancelado la transacción.')
            t = Transaccion.objects.filter(id=request.session.get('transaccion_id')).first()
            if t:
                t.estado = 'Cancelada'
                t.razon = 'Usuario canceló la transacción en el formulario'
                t.save()
                # Limpiar datos de sesión
            if 'venta_datos' in request.session:
                del request.session['venta_datos']
            if 'precio_compra_inicial' in request.session:
                del request.session['precio_compra_inicial']
            if 'transaccion_id' in request.session:
                del request.session['transaccion_id']
            return redirect('inicio')
        elif action == 'aceptar':
            transaccion = Transaccion.objects.filter(id=request.session.get('transaccion_id')).first()
            precios_actuales = transaccion.moneda.get_precios_cliente(request.user.cliente_activo)
            request.session['precio_compra_inicial'] = precios_actuales['precio_compra']
            messages.success(request, 'Precios actualizados. Continuando con la transacción.')
            context = {
                'transaccion': transaccion,
                'paso_actual': 4,
                'titulo_paso': 'Confirmación de Venta',
                'enable_2fa': is_2fa_enabled(),
                'user_email': request.user.email,
                'has_email': bool(request.user.email)
            }
            
            return render(request, 'transacciones/confirmacion.html', context)
        elif action == 'cancelar':
            t = Transaccion.objects.filter(id=request.session.get('transaccion_id')).first()
            if t:
                t.estado = 'Cancelada'
                t.razon = 'Usuario cancela debido a cambios de cotización'
                t.save()
                # Limpiar datos de sesión
            if 'venta_datos' in request.session:
                del request.session['venta_datos']
            if 'precio_compra_inicial' in request.session:
                del request.session['precio_compra_inicial']
            if 'transaccion_id' in request.session:
                del request.session['transaccion_id']
            messages.info(request, 'Transacción cancelada debido a cambios en la cotización.')
            return redirect('inicio')
    else:
        # Recuperar los datos de la sesión
        try:
            moneda = Moneda.objects.get(id=venta_datos['moneda'])
            monto = Decimal(venta_datos['monto'])
            medio_pago = venta_datos['medio_pago']
            medio_cobro = venta_datos['medio_cobro']
        except (Moneda.DoesNotExist, ValueError, KeyError):
            messages.error(request, 'Error al recuperar los datos. Reinicie el proceso.')
            return redirect('transacciones:venta_monto_moneda')
        if request.session.get('transaccion_id'):
            transaccion = Transaccion.objects.filter(id=request.session.get('transaccion_id')).first()
            context = {
            'transaccion': transaccion,
            'paso_actual': 4,
            'titulo_paso': 'Confirmación de Venta',
            'enable_2fa': is_2fa_enabled(),
            'user_email': request.user.email,
            'has_email': bool(request.user.email)
            }
            return render(request, 'transacciones/confirmacion.html', context)
    
    # Crear la transacción en la base de datos
        try:
            if medio_pago.startswith('{'):
                medio_pago_dict = ast.literal_eval(medio_pago)
                str_medio_pago = f'Tarjeta de Crédito (**** **** **** {medio_pago_dict["last4"]})'
            else:
                str_medio_pago = medio_pago
            if medio_cobro.startswith('{'):
                medio_cobro_dict = ast.literal_eval(medio_cobro)
                str_medio_cobro = f'{medio_cobro_dict["tipo_billetera"]} ({medio_cobro_dict["telefono"]})'
            else:
                str_medio_cobro = medio_cobro
                
            datos_transaccion = calcular_conversion(monto, moneda, 'venta', medio_pago, medio_cobro, request.user.cliente_activo.segmento)
            if datos_transaccion['precio_final'] > (LimiteGlobal.objects.first().limite_diario - request.user.cliente_activo.consumo_diario) or datos_transaccion['precio_final'] > (LimiteGlobal.objects.first().limite_mensual - request.user.cliente_activo.consumo_mensual):
                messages.warning(request, 'El monto final excede sus límites diarios o mensuales. Reinicie el proceso con un monto menor.')
                return redirect('transacciones:venta_monto_moneda')
            if str_medio_cobro != 'Efectivo':
                if datos_transaccion['precio_final'] > StockGuaranies.objects.first().cantidad:
                    messages.warning(request, 'El monto a recibir excede la disponibilidad actual de guaraníes en el sistema. Reinicie el proceso con un monto menor o diferente medio de cobro.')
                    return redirect('transacciones:venta_monto_moneda')
            transaccion = Transaccion.objects.create(
                cliente=request.user.cliente_activo,
                tipo='venta',
                moneda=moneda,
                monto=datos_transaccion['monto'],
                monto_original=datos_transaccion['monto_original'],
                cotizacion=datos_transaccion['cotizacion'],
                precio_base=datos_transaccion['precio_base'],
                beneficio_segmento=datos_transaccion['beneficio_segmento'],
                porc_beneficio_segmento=datos_transaccion['porc_beneficio_segmento'],
                recargo_pago=datos_transaccion['monto_recargo_pago'],
                porc_recargo_pago=datos_transaccion['porc_recargo_pago'],
                recargo_cobro=datos_transaccion['monto_recargo_cobro'],
                porc_recargo_cobro=datos_transaccion['porc_recargo_cobro'],
                redondeo_efectivo_monto=datos_transaccion['redondeo_efectivo_monto'],
                redondeo_efectivo_precio_final=datos_transaccion['redondeo_efectivo_precio_final'],
                precio_final=datos_transaccion['precio_final'],
                medio_pago=str_medio_pago,
                medio_cobro=str_medio_cobro,
                usuario=request.user
            )
            request.session['transaccion_id'] = transaccion.id
                
        except Exception as e:
            messages.error(request, 'Error al crear la transacción. Intente nuevamente.')
            return redirect('transacciones:venta_medio_cobro')
        
        cambios = verificar_cambio_cotizacion_sesion(request, 'venta')
        if cambios and cambios.get('hay_cambios'):
            transaccion = Transaccion.objects.filter(id=request.session.get('transaccion_id')).first()
            context = {
                'cambios': cambios,
                'transaccion': transaccion,
                'paso_actual': 4,
                'titulo_paso': 'Confirmación de Venta',
                'enable_2fa': is_2fa_enabled(),
                'user_email': request.user.email,
                'has_email': bool(request.user.email)
            }
            
            return render(request, 'transacciones/confirmacion.html', context)
        
        context = {
            'transaccion': transaccion,
            'paso_actual': 4,
            'titulo_paso': 'Confirmación de Venta',
            'enable_2fa': is_2fa_enabled(),
            'user_email': request.user.email,
            'has_email': bool(request.user.email)
        }
        
        return render(request, 'transacciones/confirmacion.html', context)

@login_required
def venta_exito(request):
    """
    Página final del proceso de venta: mensaje de éxito.
    
    Verifica si hay cambios en la cotización antes de finalizar la transacción.
    Si hay cambios, muestra el modal de confirmación. Si no hay cambios o el usuario
    acepta los nuevos precios, muestra confirmación de éxito.
    
    Args:
        request (HttpRequest): Petición HTTP
        
    Returns:
        HttpResponse: Página de éxito o modal de cambio de cotización
        
    Template:
        transacciones/exito.html o transacciones/cotizacion_cambiada.html
    """
    # Verificar que el usuario tenga un cliente activo
    if not request.user.cliente_activo:
        messages.error(request, 'Debe tener un cliente activo seleccionado para continuar.')
        return redirect('inicio')
    try:
        idt = request.session.get('transaccion_id')
        transaccion = Transaccion.objects.get(id=idt)
    except:
        return redirect('inicio')

    venta_datos = request.session.get('venta_datos')
    if not venta_datos:
        messages.error(request, 'Debe completar el cuarto paso antes de continuar.')
        return redirect('transacciones:venta_medio_cobro')
    try:
        medio_pago = venta_datos['medio_pago']
        medio_cobro = venta_datos.get('medio_cobro')
    except (Moneda.DoesNotExist, ValueError, KeyError):
        messages.error(request, 'Error al recuperar los datos. Reinicie el proceso.')
        return redirect('transacciones:venta_monto_moneda')
    
    if transaccion.token is None:
        if medio_pago.startswith('{'):
            medio_pago_dict = ast.literal_eval(medio_pago)
            if procesar_pago_stripe(transaccion, medio_pago_dict["id"])['success']:
                messages.success(request, 'Pago con tarjeta de crédito procesado exitosamente.')
                procesar_transaccion(transaccion)
                if medio_cobro == 'Efectivo':
                    generar_token_transaccion(transaccion)
            else:
                transaccion.estado = 'Cancelada'
                transaccion.razon = 'Error en la transacción automática de venta'
                transaccion.save()
                messages.error(request, 'Hubo un error de pago o de cobro automático. Intente nuevamente.')
                return redirect('transacciones:venta_monto_moneda')
        else:
            try:
                generar_token_transaccion(transaccion)
            except Exception as e:
                messages.error(request, 'Error al generar token de transacción. Intente nuevamente.')
                return redirect('transacciones:venta_medio_cobro')
            
    del request.session['transaccion_id']
    context = {
        'transaccion': transaccion
    }
    
    return render(request, 'transacciones/exito.html', context)

# ============================================================================
# HISTORIAL Y CONSULTA DE TRANSACCIONES
# ============================================================================

@login_required
def historial_transacciones(request, cliente_id):
    """
    Vista principal para el historial y consulta de transacciones.
    
    Muestra un listado completo de transacciones con capacidades de filtrado
    avanzado por cliente, usuario, tipo de operación y estado. Si se proporciona
    un cliente_id específico, filtra solo las transacciones de ese cliente.
    
    Filtros disponibles:
        - Búsqueda por nombre/documento de cliente o usuario
        - Tipo de operación (compra/venta)
        - Estado de transacción (pendiente/completada/etc.)
        - Usuario específico que procesó la transacción
    
    Args:
        request (HttpRequest): Petición HTTP con parámetros de filtro
        cliente_id (int, optional): ID específico de cliente para filtrar
        
    Returns:
        HttpResponse: Listado de transacciones filtrado
        
    Template:
        transacciones/historial_transacciones.html
        
    Context:
        - transacciones: QuerySet de transacciones filtradas
        - busqueda: Término de búsqueda aplicado
        - tipo_operacion: Filtro de tipo aplicado
        - estado_filtro: Filtro de estado aplicado
        - cliente_filtrado: Cliente específico si aplica
        - usuario_filtro: Usuario específico si aplica
        - usuarios_cliente: Usuarios asociados al cliente (si aplica)
    """
    try:
        cliente = Cliente.objects.get(id=cliente_id)
        transacciones = Transaccion.objects.filter(cliente=cliente).order_by('-fecha_hora')
    except Cliente.DoesNotExist:
        messages.error(request, "Cliente no encontrado")
        return redirect('transacciones:historial')
    if not request.user.clientes_operados.filter(id=cliente.id).exists():
        messages.error(request, "No tienes permiso para ver el historial de este cliente.")
        return redirect('inicio')
    transacciones_pasadas = Transaccion.objects.filter(cliente=cliente, estado='Pendiente')
    if transacciones_pasadas:
        for t in transacciones_pasadas:
            if t.fecha_hora < timezone.now() - timedelta(minutes=5):
                t.estado = 'Cancelada'
                t.razon = 'Expira el tiempo para confirmar la transacción'
                t.save()
    # Obtener parámetros de filtrado
    busqueda = request.GET.get('busqueda', '')
    tipo_operacion = request.GET.get('tipo_operacion', '')
    estado_filtro = request.GET.get('estado', '')
    usuario_filtro = request.GET.get('usuario', '')
    
    # Aplicar filtros según parámetros recibidos
    if busqueda:
        # Buscar por cliente o usuario solo cuando no hay cliente filtrado
        transacciones = transacciones.filter(
            models.Q(cliente__nombre__icontains=busqueda) | 
            models.Q(cliente__numero_documento__icontains=busqueda) |
            models.Q(cliente__usuarios__username__icontains=busqueda)
        )
    
    if tipo_operacion:
        transacciones = transacciones.filter(tipo=tipo_operacion)
    
    if estado_filtro:
        transacciones = transacciones.filter(estado__iexact=estado_filtro)
    
    # Filtrar por usuario si se especifica
    if usuario_filtro:
        try:
            usuario_id = int(usuario_filtro)
            transacciones = transacciones.filter(usuario_id=usuario_id)
        except (ValueError, TypeError):
            pass
    
    context = {
        'transacciones': transacciones,
        'busqueda': busqueda,
        'tipo_operacion': tipo_operacion,
        'estado_filtro': estado_filtro,
        'usuario_filtro': usuario_filtro,
        'cliente_filtrado': cliente,
        'usuarios_cliente': cliente.usuarios.all(),
    }
    
    return render(request, 'transacciones/historial_transacciones.html', context)

@login_required
def detalle_historial(request, transaccion_id):
    """
    Vista de detalle para una transacción específica.
    
    Muestra información completa y detallada de una transacción individual,
    incluyendo todos los datos relevantes como montos calculados, medios
    de pago/cobro, información del cliente y estado actual.
    
    Cálculos realizados:
        - Para compras: Convierte monto extranjero a guaraníes usando precio de venta
        - Para ventas: Convierte monto extranjero a guaraníes usando precio de compra
        - Aplica beneficios del segmento del cliente
    
    Args:
        request (HttpRequest): Petición HTTP
        transaccion_id (int): ID de la transacción a mostrar
        
    Returns:
        HttpResponse: Página de detalle o redirecciona si no existe
        
    Template:
        transacciones/detalle_transaccion.html
        
    Context:
        - transaccion: Instancia de la transacción
        - monto_origen: Monto en moneda de origen (guaraníes o extranjera)
        - monto_destino: Monto en moneda de destino (extranjera o guaraníes)
    """
    try:
        transaccion = Transaccion.objects.get(id=transaccion_id)
    except Transaccion.DoesNotExist:
        messages.error(request, 'La transacción solicitada no existe.')
        return redirect('transacciones:historial')
    transacciones_pasadas = Transaccion.objects.filter(cliente=transaccion.cliente, estado='Pendiente')
    if transacciones_pasadas:
        for t in transacciones_pasadas:
            if t.fecha_hora < timezone.now() - timedelta(minutes=5):
                t.estado = 'Cancelada'
                t.razon = 'Expira el tiempo para confirmar la transacción'
                t.save()
    if transaccion.cliente not in request.user.clientes_operados.all():
        messages.error(request, "No tienes permiso para ver el historial de este cliente.")
        return redirect('inicio')
    
    context = {
        'transaccion': transaccion,
    }
    
    return render(request, 'transacciones/detalle_historial.html', context)

@login_required
@permission_required('transacciones.edicion', raise_exception=True)
def ver_variables(request):
    """
    Vista para ver variables de gestión de transacciones.
    """
    context = {
        'limites': LimiteGlobal.objects.first(),
        'recargos': Recargos.objects.all()
    }
    return render(request, 'transacciones/ver_variables.html', context)

@login_required
@permission_required('transacciones.edicion', raise_exception=True)
def editar_variables(request):
    """
    Vista para editar variables de gestión de transacciones.
    """
    if request.method == 'POST':
        form = VariablesForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Variables actualizadas correctamente.')
            return redirect('transacciones:ver_variables')
    else:
        form = VariablesForm()

    context = {
        'form': form
    }
    return render(request, 'transacciones/editar_variables.html', context)

@login_required
@permission_required('transacciones.revision', raise_exception=True)
def revisar_tausers(request):
    """
    Vista para revisar la información de los TAUsers.
    
    Muestra un listado de todos los TAUsers del sistema con su información básica
    incluyendo puerto, estado de activación (detectado dinámicamente) y opciones de acción.
    
    Funcionalidades:
        - Listado completo de TAUsers
        - Búsqueda por puerto
        - Filtrado por estado (activo/inactivo) detectado automáticamente
        - Estadísticas de TAUsers activos e inactivos
        - Opción de ver detalles de cada TAUser
    
    Args:
        request (HttpRequest): Petición HTTP con parámetros de filtro opcionales
        
    Returns:
        HttpResponse: Renderiza listado de TAUsers
        
    Template:
        transacciones/tausers_lista.html
        
    Context:
        - tausers: Lista de TAUsers con estado detectado dinámicamente
        - busqueda: Término de búsqueda aplicado
        - tausers_activos: Cantidad de TAUsers activos
        - tausers_inactivos: Cantidad de TAUsers inactivos
        - total_tausers_sistema: Total de TAUsers en el sistema
        - filtro_estado: Estado seleccionado para filtrar
    """
    # Obtener todos los TAUsers
    todos_los_tausers = Tauser.objects.all()
    tausers_query = todos_los_tausers
    
    # Manejar búsqueda por sucursal
    busqueda = request.GET.get('busqueda', '').strip()
    if busqueda:
        tausers_query = tausers_query.filter(sucursal__icontains=busqueda)
    
    # Ordenar por número de Tauser
    tausers_query = tausers_query.order_by('id')
    
    # Obtener filtro de estado
    filtro_estado = request.GET.get('estado', '')
    
    # Convertir QuerySet a lista y agregar estado dinámico
    tausers_con_estado = []
    tausers_activos_count = 0
    tausers_inactivos_count = 0
    
    for tauser in tausers_query:
        # Detectar estado dinámicamente
        esta_activo = tauser.esta_activo()
        
        # Aplicar filtro de estado si se especifica
        if filtro_estado == 'activo' and not esta_activo:
            continue
        elif filtro_estado == 'inactivo' and esta_activo:
            continue
        
        # Agregar el tauser con su estado detectado
        tauser.activo = esta_activo  # Agregar atributo temporal
        tausers_con_estado.append(tauser)
        
        # Contar para estadísticas
        if esta_activo:
            tausers_activos_count += 1
        else:
            tausers_inactivos_count += 1
    
    # Total de TAUsers en el sistema (sin filtrar)
    total_tausers_sistema = todos_los_tausers.count()
    
    context = {
        'tausers': tausers_con_estado,
        'tausers_activos': tausers_activos_count,
        'tausers_inactivos': tausers_inactivos_count,
        'busqueda': busqueda,
        'filtro_estado': filtro_estado,
        'total_tausers_sistema': total_tausers_sistema,
    }
    return render(request, 'transacciones/tausers_lista.html', context)

@login_required
@permission_required('transacciones.revision', raise_exception=True)
def tauser_detalle(request, pk):
    """
    Vista para mostrar los detalles completos de un TAUser específico.
    
    Muestra información detallada del TAUser incluyendo puerto, estado,
    billetes asociados y sus cantidades disponibles.
    Incluye filtrado por moneda para billetes.
    
    Args:
        request (HttpRequest): Petición HTTP
        pk (int): ID del TAUser a mostrar
        
    Returns:
        HttpResponse: Página de detalle del TAUser
        
    Template:
        transacciones/tauser_detalles.html
        
    Context:
        - tauser: Instancia del TAUser
        - billetes_tauser: QuerySet de billetes asociados al TAUser
        - monedas_disponibles: Lista de monedas que tienen billetes en este TAUser
        - moneda_filtro: Moneda seleccionada para filtrar (si aplica)  
        - totales_por_moneda: Diccionario con totales por moneda
    """
    from .models import BilletesTauser
    from monedas.models import Moneda
    from django.db.models import Sum, F, Count
    
    tauser = get_object_or_404(Tauser, pk=pk)
    
    # Obtener todos los billetes del TAUser (sin filtrar para obtener todas las monedas disponibles)
    todos_billetes_tauser = BilletesTauser.objects.filter(tauser=tauser)
    
    # Calcular totales por moneda
    totales_por_moneda = {}
    
    # Total para guaraníes (denominaciones sin moneda)
    total_guaranies = todos_billetes_tauser.filter(
        denominacion__moneda__isnull=True
    ).aggregate(
        total=Sum(F('cantidad') * F('denominacion__valor'))
    )['total'] or 0
    
    if total_guaranies > 0:
        totales_por_moneda['guarani'] = {
            'nombre': 'Guaraní',
            'total': total_guaranies,
            'simbolo': 'Gs.'
        }
    
    # Totales para monedas extranjeras
    for moneda in Moneda.objects.all():
        total_moneda = todos_billetes_tauser.filter(
            denominacion__moneda=moneda
        ).aggregate(
            total=Sum(F('cantidad') * F('denominacion__valor'))
        )['total']
        
        if total_moneda and total_moneda > 0:
            totales_por_moneda[moneda.id] = {
                'nombre': moneda.nombre,
                'total': total_moneda,
                'simbolo': moneda.simbolo
            }
    
    # Obtener todas las monedas que tienen billetes en este TAUser (antes de aplicar filtros)
    monedas_con_billetes = todos_billetes_tauser.values_list('denominacion__moneda', flat=True).distinct()
    monedas_disponibles = []
    
    # Agregar guaraní si hay billetes sin moneda (denominación de guaraníes)
    if None in monedas_con_billetes:
        monedas_disponibles.append({'id': 'guarani', 'nombre': 'Guaraní'})
    
    # Agregar monedas extranjeras
    for moneda_id in monedas_con_billetes:
        if moneda_id is not None:
            try:
                moneda = Moneda.objects.get(id=moneda_id)
                monedas_disponibles.append({'id': moneda.id, 'nombre': moneda.nombre})
            except Moneda.DoesNotExist:
                pass
    
    # Ahora aplicar el filtro para mostrar los billetes
    billetes_tauser_query = todos_billetes_tauser
    
    # Obtener parámetro de filtro de moneda
    moneda_filtro = request.GET.get('moneda', '')
    
    # Aplicar filtro de moneda si se especifica
    if moneda_filtro:
        if moneda_filtro == 'guarani':
            billetes_tauser_query = billetes_tauser_query.filter(denominacion__moneda__isnull=True)
        else:
            try:
                moneda_id = int(moneda_filtro)
                billetes_tauser_query = billetes_tauser_query.filter(denominacion__moneda_id=moneda_id)
            except (ValueError, TypeError):
                pass
    
    billetes_tauser = billetes_tauser_query.order_by('denominacion__moneda', 'denominacion__valor')
    
    context = {
        'tauser': tauser,
        'billetes_tauser': billetes_tauser,
        'monedas_disponibles': monedas_disponibles,
        'moneda_filtro': moneda_filtro,
        'totales_por_moneda': totales_por_moneda
    }
    return render(request, 'transacciones/tauser_detalles.html', context)

@login_required
@permission_required('transacciones.revision', raise_exception=True) 
def verificar_estado_tauser(request, tauser_id):
    """
    Vista AJAX para verificar el estado de un TAUser específico.
    
    Útil para verificaciones individuales sin recargar toda la página.
    
    Args:
        request (HttpRequest): Petición HTTP (debe ser AJAX)
        tauser_id (int): ID del TAUser a verificar
        
    Returns:
        JsonResponse: Estado del TAUser y información adicional
    """
    from django.http import JsonResponse
    
    try:
        tauser = get_object_or_404(Tauser, pk=tauser_id)
        esta_activo = tauser.esta_activo()
        
        return JsonResponse({
            'success': True,
            'tauser_id': tauser.id,
            'puerto': tauser.puerto,
            'activo': esta_activo,
            'estado_texto': 'Activo' if esta_activo else 'Inactivo',
            'url': f'http://localhost:{tauser.puerto}'
        })
        
    except Tauser.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'TAUser no encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json

logger = logging.getLogger(__name__)

@csrf_exempt
def recibir_pago(request):
    """Vista para recibir y verificar notificaciones de la pasarela"""
    logger.info("Recibiendo notificación de pago")
    
    if request.method == 'POST':
        try:
            # Log de headers para debug
            logger.info(f"Headers recibidos: {dict(request.headers)}")
            

            data = json.loads(request.body)
            logger.info(f"Datos recibidos: {data}")
            
            estado = data.get('estado')
            monto = data.get('monto')
            tipo_pago = data.get('tipo_pago')
            
            # Validar datos
            if estado not in ['exito', 'error', 'cancelado']:
                logger.error(f"Estado inválido recibido: {estado}")
                return JsonResponse({'status': 'error', 'mensaje': 'Estado inválido'})
            
            if estado != 'cancelado' and not monto:
                logger.error("Monto requerido no recibido")
                return JsonResponse({'status': 'error', 'mensaje': 'Monto requerido'})

            # Procesar según estado
            if estado == 'exito':
                logger.info(f"Pago exitoso procesado: {data}")
                return JsonResponse({'status': 'ok', 'mensaje': 'Pago procesado correctamente'})
            elif estado == 'error':
                logger.warning(f"Error en pago: {data}")
                return JsonResponse({'status': 'error', 'mensaje': 'Error procesando pago'})
            else:
                logger.info(f"Pago cancelado: {data}")
                return JsonResponse({'status': 'ok', 'mensaje': 'Pago cancelado'})

        except json.JSONDecodeError as e:
            logger.error(f"Error decodificando JSON: {str(e)}")
            return JsonResponse({'status': 'error', 'mensaje': 'JSON inválido'}, status=400)
        except Exception as e:
            logger.error(f"Error inesperado: {str(e)}")
            return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=500)
    
    logger.warning("Método no permitido")
    return JsonResponse({'status': 'error', 'mensaje': 'Método no permitido'}, status=405)



# ============================================================================
# VISTAS PARA DESCARGA DE HISTORIAL DE TRANSACCIONES
# ============================================================================

@login_required
def descargar_historial_pdf(request):
    """
    Vista para descargar el historial de transacciones en formato PDF.
    
    Genera un archivo PDF profesional con las transacciones filtradas, donde cada 
    transacción se presenta en su propia página individual con títulos, filtros 
    aplicados y detalles completos de la operación.
    
    Características del PDF generado:
        - Formato A4 con márgenes de 2cm en todos los bordes
        - Cada página contiene: títulos corporativos, filtros aplicados y una transacción
        - Tablas optimizadas con fuentes legibles y espaciado profesional
        - Información completa de cada transacción incluyendo cálculos y estados
        - Formato de fecha/hora sincronizado con la zona horaria local
    
    Filtros soportados:
        - cliente: ID del cliente específico (requerido)
        - busqueda: Búsqueda por nombre/documento de cliente o username de usuario
        - tipo_operacion: Filtro por tipo ('compra' o 'venta')
        - estado: Filtro por estado de transacción
        - usuario: ID del usuario que procesó la transacción
    
    Seguridad:
        - Verifica permisos del usuario para acceder al cliente solicitado
        - Valida existencia del cliente antes de procesar
        - Aplica los mismos filtros de seguridad que la vista de historial
    
    Estructura del PDF:
        - Página 1: Títulos + Filtros + Primera transacción
        - Páginas N: Títulos + Filtros + Transacción N
        - Cada transacción incluye: tabla principal y tabla de detalles expandida
    
    Args:
        request (HttpRequest): Petición HTTP con parámetros de filtro en GET
            - cliente (str): ID del cliente (requerido)
            - busqueda (str, opcional): Término de búsqueda
            - tipo_operacion (str, opcional): 'compra' o 'venta'
            - estado (str, opcional): Estado de la transacción
            - usuario (str, opcional): ID del usuario
        
    Returns:
        HttpResponse: Archivo PDF para descarga con content-type 'application/pdf'
            - Nombre del archivo: 'historial_transacciones_{cliente_nombre}.pdf'
            - Headers configurados para descarga automática
            
    Raises:
        HttpResponse(400): Si no se proporciona ID de cliente
        HttpResponse(404): Si el cliente no existe
        redirect('inicio'): Si el usuario no tiene permisos para el cliente
    
    Dependencias:
        - ReportLab: Para generación de PDF (SimpleDocTemplate, Table, etc.)
        - Django timezone: Para formateo de fechas consistente
        - Modelos: Cliente, Transaccion, Usuario
    
    """
    from django.http import HttpResponse
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from io import BytesIO
    
    # Obtener parámetros de filtro
    cliente_id = request.GET.get('cliente')
    busqueda = request.GET.get('busqueda', '')
    tipo_operacion = request.GET.get('tipo_operacion', '')
    estado_filtro = request.GET.get('estado', '')
    usuario_filtro = request.GET.get('usuario', '')
    
    # Obtener transacciones con los mismos filtros que la vista principal
    try:
        if cliente_id:
            cliente = Cliente.objects.get(id=cliente_id)
            transacciones = Transaccion.objects.filter(cliente=cliente).order_by('-fecha_hora')
            
            # Verificar permisos del usuario
            if not request.user.clientes_operados.filter(id=cliente.id).exists():
                messages.error(request, "No tienes permiso para descargar el historial de este cliente.")
                return redirect('inicio')
        else:
            return HttpResponse("Cliente no especificado", status=400)
    except Cliente.DoesNotExist:
        return HttpResponse("Cliente no encontrado", status=404)
    
    # Aplicar los mismos filtros que en la vista de historial
    if busqueda:
        transacciones = transacciones.filter(
            models.Q(cliente__nombre__icontains=busqueda) | 
            models.Q(cliente__numero_documento__icontains=busqueda) |
            models.Q(cliente__usuarios__username__icontains=busqueda)
        )
    
    if tipo_operacion:
        transacciones = transacciones.filter(tipo=tipo_operacion)
    
    if estado_filtro:
        transacciones = transacciones.filter(estado__iexact=estado_filtro)
    
    if usuario_filtro:
        try:
            usuario_id = int(usuario_filtro)
            transacciones = transacciones.filter(usuario_id=usuario_id)
        except (ValueError, TypeError):
            pass
    
    # Crear el PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    
    # Crear sección de transacciones con detalles
    for i, transaccion in enumerate(transacciones):
        # Si no es la primera transacción, agregar salto de página
        if i > 0:
            story.append(PageBreak())
        
        # Agregar títulos y filtros en cada página
        # Título de la empresa
        company_title = "Global Exchange"
        title_style = styles['Title'].clone('CompactTitle')
        title_style.fontSize = 20  # Ligeramente aumentado
        title_style.leading = 24
        story.append(Paragraph(company_title, title_style))
        story.append(Spacer(1, 8))  # Ligeramente aumentado
        
        # Título del historial centrado
        historial_title = f"Historial de Transacciones - {cliente.nombre}"
        # Crear estilo centrado para el título
        centered_style = styles['Heading1'].clone('CenteredHeading1')
        centered_style.fontSize = 15  # Ligeramente aumentado
        centered_style.leading = 18
        centered_style.alignment = 1  # 1 = CENTER
        story.append(Paragraph(historial_title, centered_style))
        story.append(Spacer(1, 12))  # Ligeramente aumentado
        
        # Información de filtros aplicados
        if busqueda or tipo_operacion or estado_filtro or usuario_filtro:
            filtro_style = styles['Normal'].clone('CompactNormal')
            filtro_style.fontSize = 11  # Ligeramente aumentado
            filtro_style.leading = 13
            story.append(Paragraph("Filtros aplicados:", filtro_style))
            story.append(Spacer(1, 5))  # Ligeramente aumentado
            
            if busqueda:
                story.append(Paragraph(f"• Búsqueda: '{busqueda}'", filtro_style))
            if tipo_operacion:
                story.append(Paragraph(f"• Tipo de operación: {tipo_operacion.title()}", filtro_style))
            if estado_filtro:
                story.append(Paragraph(f"• Estado: {estado_filtro.title()}", filtro_style))
            if usuario_filtro:
                try:
                    from usuarios.models import Usuario
                    usuario = Usuario.objects.get(id=usuario_filtro)
                    story.append(Paragraph(f"• Usuario: {usuario.nombre_completo() or usuario.username}", filtro_style))
                except:
                    pass
            
            story.append(Spacer(1, 10))  # Ligeramente aumentado
        
        # Crear tabla principal para cada transacción
        # Usar timezone local como en el template HTML
        fecha_hora_local = timezone.localtime(transaccion.fecha_hora)
        fecha_str = fecha_hora_local.strftime("%d/%m/%Y %H:%M:%S")
        usuario_str = transaccion.usuario.nombre_completo() or transaccion.usuario.username
        operacion_str = transaccion.tipo.title()
        # Formatear el monto con los decimales correspondientes a la moneda
        monto_formateado = f"{transaccion.monto:.{transaccion.moneda.decimales}f}"
        moneda_str = f"{monto_formateado} {transaccion.moneda.simbolo}"
        estado_str = transaccion.estado
        
        # Cada transacción tiene su propio encabezado
        moneda_simbolo = transaccion.moneda.simbolo  # Solo el símbolo sin monto
        main_data = [
            ['Fecha/Hora', 'Operación', 'Moneda', 'Estado'],
            [fecha_str, operacion_str, moneda_simbolo, estado_str]
        ]
        
        
        # Distribuir proporcionalmente: Fecha/Hora más ancha, otras iguales
        main_colWidths = [2.4*inch, 1.15*inch, 1.15*inch, 1.15*inch]  # Total: 5.85 inches (ligeramente aumentado)
        main_table = Table(main_data, colWidths=main_colWidths)
        main_table.setStyle(TableStyle([
            # Estilo del encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),  # Ligeramente aumentado
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Ligeramente aumentado
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            # Estilo de la fila de datos
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 11),  # Ligeramente aumentado
            ('BOTTOMPADDING', (0, 1), (-1, -1), 10),  # Ligeramente aumentado
            ('TOPPADDING', (0, 1), (-1, -1), 7),
            # Alineación y bordes
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(main_table)
        
        # Tabla de detalles completos de la transacción
        details_data = [
            ['Detalles Completos de la Transacción'],
            ['Usuario', usuario_str],
            ['Segmento Cliente', transaccion.cliente.get_segmento_display()],
            ['Tipo de Transacción', f"{transaccion.tipo.title()} de {transaccion.moneda.nombre}"],
            ['Medio de Pago', transaccion.medio_pago],
            ['Medio de Cobro', transaccion.medio_cobro]
        ]
        
        # Detalles específicos por tipo de transacción
        monto_original_formateado = f"{transaccion.monto_original:.{transaccion.moneda.decimales}f}"
        monto_formateado = f"{transaccion.monto:.{transaccion.moneda.decimales}f}"
        redondeo_monto_formateado = f"{transaccion.redondeo_efectivo_monto:.{transaccion.moneda.decimales}f}"
        
        if transaccion.tipo == 'compra':
            details_data.extend([
                ['Monto ingresado para compra', f"{monto_original_formateado} {transaccion.moneda.simbolo}"],
                ['Redondeo por medio cobro Efectivo', f"{redondeo_monto_formateado} {transaccion.moneda.simbolo}"],
                ['Monto a comprar', f"{monto_formateado} {transaccion.moneda.simbolo}"],
                ['Cotización para compra', f"Gs. {transaccion.cotizacion:,.0f}"],
                ['Precio base para compra', f"Gs. {transaccion.precio_base:,.0f}"]
            ])
        else:
            details_data.extend([
                ['Monto ingresado para venta', f"{monto_original_formateado} {transaccion.moneda.simbolo}"],
                ['Monto a vender', f"{monto_formateado} {transaccion.moneda.simbolo}"],
                ['Cotización para venta', f"Gs. {transaccion.cotizacion:,.0f}"],
                ['Precio base para venta', f"Gs. {transaccion.precio_base:,.0f}"]
            ])
            
            if transaccion.medio_pago == 'Efectivo':
                details_data.append(['Redondeo por medio pago Efectivo', f"{redondeo_monto_formateado} {transaccion.moneda.simbolo}"])
        
        # Beneficios y recargos
        if transaccion.beneficio_segmento and transaccion.beneficio_segmento > 0:
            details_data.append(['Beneficio por segmento', f"Gs. {transaccion.beneficio_segmento:,.0f}"])
        
        
        details_data.append([f'Recargo por medio de pago ({transaccion.porc_recargo_pago})', f"Gs. {transaccion.recargo_pago:,.0f}"])
        
        
        if transaccion.tipo == 'venta':
            details_data.append([f'Recargo por medio de cobro ({transaccion.porc_recargo_cobro})', f"Gs. {transaccion.recargo_cobro:,.0f}"])
        
        # Redondeo efectivo para precio final si aplica
        if ((transaccion.tipo == 'compra' and transaccion.medio_pago == 'Efectivo') or 
            (transaccion.tipo == 'venta' and transaccion.medio_cobro == 'Efectivo')):
            details_data.append(['Redondeo efectivo precio final', f"Gs. {transaccion.redondeo_efectivo_precio_final:,.0f}"])
        
        # Montos finales
        if transaccion.tipo == 'compra':
            estado_texto = 'pagado' if transaccion.estado in ['Completa', 'Confirmada'] else 'a pagar'
            recibido_texto = 'recibido' if transaccion.estado == 'Completa' else 'a recibir'
            
            details_data.extend([
                [f'Monto {estado_texto}', f"Gs. {transaccion.precio_final:,.0f}"],
                [f'Monto {recibido_texto}', f"{monto_formateado} {transaccion.moneda.simbolo}"]
            ])
        else:
            estado_texto = 'pagado' if transaccion.estado in ['Completa', 'Confirmada'] else 'a pagar'
            recibido_texto = 'recibido' if transaccion.estado == 'Completa' else 'a recibir'
            
            details_data.extend([
                [f'Monto {estado_texto}', f"{monto_formateado} {transaccion.moneda.simbolo}"],
                [f'Monto {recibido_texto}', f"Gs. {transaccion.precio_final:,.0f}"]
            ])
        
        # Mostrar monto pagado si es parcial
        if ((transaccion.tipo == 'compra' and transaccion.pagado < transaccion.precio_final and transaccion.estado != 'Completa') or
            (transaccion.tipo == 'venta' and transaccion.pagado < transaccion.monto and transaccion.estado != 'Completa')):
            if transaccion.tipo == 'compra':
                details_data.append(['Monto pagado', f"Gs. {transaccion.pagado:,.0f}"])
            else:
                details_data.append(['Monto pagado', f"{transaccion.pagado:.{transaccion.moneda.decimales}f} {transaccion.moneda.simbolo}"])
        
        # Token si existe y está pendiente/confirmada
        if transaccion.token and transaccion.estado in ['Pendiente', 'Confirmada']:
            details_data.append(['Código de transacción', transaccion.token])
        
        details_table = Table(details_data, colWidths=[2.4*inch, 3.45*inch])  # Ligeramente aumentado
        details_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),  # Ligeramente aumentado
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),  # Ligeramente aumentado
            ('TOPPADDING', (0, 0), (-1, 0), 7),
            ('SPAN', (0, 0), (-1, 0)),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),  # Ligeramente aumentado
            ('BOTTOMPADDING', (0, 1), (-1, -1), 7),  # Ligeramente aumentado
            ('TOPPADDING', (0, 1), (-1, -1), 5),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('LEFTPADDING', (0, 1), (-1, -1), 7),  # Ligeramente aumentado
            ('RIGHTPADDING', (0, 1), (-1, -1), 7)  # Ligeramente aumentado
        ]))
        
        story.append(details_table)
        story.append(Spacer(1, 15))  # Espacio ligeramente mayor al final
    
    # Construir el PDF
    doc.build(story)
    
    # Preparar la respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    filename = f"historial_transacciones_{cliente.nombre.replace(' ', '_')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required  
def descargar_historial_excel(request):
    """
    Vista para descargar el historial de transacciones en formato Excel.
    
    Genera un archivo Excel (.xlsx) con formato profesional que incluye títulos 
    corporativos, filtros aplicados y detalles expandidos de cada transacción 
    con encabezados individuales para mejor organización.
    
    Características del Excel generado:
        - Hoja única con todas las transacciones organizadas secuencialmente
        - Títulos corporativos y filtros aplicados al inicio
        - Cada transacción tiene su propio encabezado con información principal
        - Detalles expandidos con todos los cálculos y montos de cada operación
        - Formato profesional con colores, fuentes en negrita y alineación
        - Anchos de columna optimizados para legibilidad
        - Bordes y estilos aplicados consistentemente
    
    Filtros soportados:
        - cliente: ID del cliente específico (requerido)
        - busqueda: Búsqueda por nombre/documento de cliente o username de usuario
        - tipo_operacion: Filtro por tipo ('compra' o 'venta')
        - estado: Filtro por estado de transacción
        - usuario: ID del usuario que procesó la transacción
    
    Seguridad:
        - Verifica permisos del usuario para acceder al cliente solicitado
        - Valida existencia del cliente antes de procesar
        - Aplica los mismos filtros de seguridad que la vista de historial
    
    Estructura del Excel:
        - Fila 1-3: Título "Global Exchange" y "Historial de Transacciones"
        - Fila 4+: Filtros aplicados (si existen)
        - Por cada transacción:
            * Encabezado con datos principales (fecha, operación, moneda, estado)
            * Detalle expandido con todos los cálculos, medios de pago/cobro
            * Montos finales y información de pagos parciales
            * Código de transacción si aplica
    
    Formato y Estilos:
        - Encabezados en negrita con fondo gris
        - Datos alineados apropiadamente (centro, izquierda, derecha)
        - Bordes en todas las celdas para mejor visualización
        - Anchos de columna: [25, 18, 20, 15, 15] para óptima legibilidad
        - Formato de números con separadores de miles
    
    Args:
        request (HttpRequest): Petición HTTP con parámetros de filtro en GET
            - cliente (str): ID del cliente (requerido)
            - busqueda (str, opcional): Término de búsqueda
            - tipo_operacion (str, opcional): 'compra' o 'venta'
            - estado (str, opcional): Estado de la transacción
            - usuario (str, opcional): ID del usuario
        
    Returns:
        HttpResponse: Archivo Excel para descarga con content-type 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            - Nombre del archivo: 'historial_transacciones_{cliente_nombre}.xlsx'
            - Headers configurados para descarga automática
            
    Raises:
        HttpResponse(400): Si no se proporciona ID de cliente
        HttpResponse(404): Si el cliente no existe
        redirect('inicio'): Si el usuario no tiene permisos para el cliente
    
    Dependencias:
        - OpenPyXL: Para generación de archivos Excel (Workbook, estilos, etc.)
        - Django timezone: Para formateo de fechas consistente
        - Modelos: Cliente, Transaccion, Usuario
    
    """
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    # Obtener parámetros de filtro
    cliente_id = request.GET.get('cliente')
    busqueda = request.GET.get('busqueda', '')
    tipo_operacion = request.GET.get('tipo_operacion', '')
    estado_filtro = request.GET.get('estado', '')
    usuario_filtro = request.GET.get('usuario', '')
    
    # Obtener transacciones con los mismos filtros que la vista principal
    try:
        if cliente_id:
            cliente = Cliente.objects.get(id=cliente_id)
            transacciones = Transaccion.objects.filter(cliente=cliente).order_by('-fecha_hora')
            
            # Verificar permisos del usuario
            if not request.user.clientes_operados.filter(id=cliente.id).exists():
                messages.error(request, "No tienes permiso para descargar el historial de este cliente.")
                return redirect('inicio')
        else:
            return HttpResponse("Cliente no especificado", status=400)
    except Cliente.DoesNotExist:
        return HttpResponse("Cliente no encontrado", status=404)
    
    # Aplicar los mismos filtros que en la vista de historial
    if busqueda:
        transacciones = transacciones.filter(
            models.Q(cliente__nombre__icontains=busqueda) | 
            models.Q(cliente__numero_documento__icontains=busqueda) |
            models.Q(cliente__usuarios__username__icontains=busqueda)
        )
    
    if tipo_operacion:
        transacciones = transacciones.filter(tipo=tipo_operacion)
    
    if estado_filtro:
        transacciones = transacciones.filter(estado__iexact=estado_filtro)
    
    if usuario_filtro:
        try:
            usuario_id = int(usuario_filtro)
            transacciones = transacciones.filter(usuario_id=usuario_id)
        except (ValueError, TypeError):
            pass
    
    # Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial de Transacciones"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    
    # Título de la empresa (fila 1)
    ws.merge_cells('A1:E1') 
    ws['A1'] = "Global Exchange"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Título del historial (fila 2)
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Historial de Transacciones - {cliente.nombre}"
    ws['A2'].font = Font(bold=True, size=14)
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Información de filtros (si aplica)
    row_start = 4  
    if busqueda or tipo_operacion or estado_filtro or usuario_filtro:
        current_row = 3  # Comenzar en la fila 3 después de los títulos
        
        # Título de filtros
        ws.merge_cells(f'A{current_row}:E{current_row}')  # Ahora son 5 columnas
        ws[f'A{current_row}'] = "Filtros aplicados:"
        ws[f'A{current_row}'].font = Font(bold=True, italic=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        current_row += 1
        
        # Filtros individuales por fila
        if busqueda:
            ws.merge_cells(f'A{current_row}:E{current_row}') 
            ws[f'A{current_row}'] = f"• Búsqueda: '{busqueda}'"
            ws[f'A{current_row}'].font = Font(italic=True)
            ws[f'A{current_row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            current_row += 1
        if tipo_operacion:
            ws.merge_cells(f'A{current_row}:E{current_row}') 
            ws[f'A{current_row}'] = f"• Tipo de operación: {tipo_operacion.title()}"
            ws[f'A{current_row}'].font = Font(italic=True)
            ws[f'A{current_row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            current_row += 1
        if estado_filtro:
            ws.merge_cells(f'A{current_row}:E{current_row}') 
            ws[f'A{current_row}'] = f"• Estado: {estado_filtro.title()}"
            ws[f'A{current_row}'].font = Font(italic=True)
            ws[f'A{current_row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            current_row += 1
        if usuario_filtro:
            try:
                from usuarios.models import Usuario
                usuario = Usuario.objects.get(id=usuario_filtro)
                ws.merge_cells(f'A{current_row}:E{current_row}')  # Ahora son 5 columnas
                ws[f'A{current_row}'] = f"• Usuario: {usuario.nombre_completo() or usuario.username}"
                ws[f'A{current_row}'].font = Font(italic=True)
                ws[f'A{current_row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                current_row += 1
            except:
                pass
        
        row_start = current_row + 1
    
    # Datos de las transacciones con detalles - cada una con su propio encabezado
    current_row = row_start
    
    for transaccion in transacciones:
        # Agregar encabezado para cada transacción
        headers = ['Fecha', 'Hora', 'Operación', 'Moneda', 'Estado']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        current_row += 1
        # Fila principal de la transacción - sin columna Usuario y Moneda solo símbolo
        # Usar timezone local como en el template HTML
        fecha_hora_local = timezone.localtime(transaccion.fecha_hora)
        data_row = [
            fecha_hora_local.strftime("%d/%m/%Y"),
            fecha_hora_local.strftime("%H:%M:%S"),
            transaccion.tipo.title(),
            transaccion.moneda.simbolo,  
            transaccion.estado
        ]
        
        # Aplicar estilo mejorado a la fila principal
        for col_num, value in enumerate(data_row, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = value
            cell.border = border
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        current_row += 1
        
        # Filas de detalles completos de la transacción
        detail_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        detail_font = Font(italic=True, size=9)
        
        def add_detail_row(label, value):
            """Helper para agregar filas de detalle con formato mejorado"""
            # Celda de etiqueta
            label_cell = ws.cell(row=current_row, column=1)
            label_cell.value = label
            label_cell.font = Font(bold=True, size=10)
            label_cell.fill = detail_fill
            label_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            label_cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Celda de valor
            value_cell = ws.cell(row=current_row, column=2)
            value_cell.value = value
            value_cell.font = detail_font
            value_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            value_cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            return current_row + 1
        
        # Información básica
        current_row = add_detail_row("Usuario:", transaccion.usuario.nombre_completo() or transaccion.usuario.username)
        current_row = add_detail_row("Segmento Cliente:", transaccion.cliente.get_segmento_display())
        current_row = add_detail_row("Tipo de Transacción:", f"{transaccion.tipo.title()} de {transaccion.moneda.nombre}")
        current_row = add_detail_row("Medio de Pago:", transaccion.medio_pago)
        current_row = add_detail_row("Medio de Cobro:", transaccion.medio_cobro)
        
        # Detalles específicos por tipo de transacción
        monto_original_formateado = f"{transaccion.monto_original:.{transaccion.moneda.decimales}f}"
        monto_formateado = f"{transaccion.monto:.{transaccion.moneda.decimales}f}"
        redondeo_monto_formateado = f"{transaccion.redondeo_efectivo_monto:.{transaccion.moneda.decimales}f}"
        
        if transaccion.tipo == 'compra':
            current_row = add_detail_row("Monto ingresado para compra:", f"{monto_original_formateado} {transaccion.moneda.simbolo}")
            current_row = add_detail_row("Redondeo por medio cobro Efectivo:", f"{redondeo_monto_formateado} {transaccion.moneda.simbolo}")
            current_row = add_detail_row("Monto a comprar:", f"{monto_formateado} {transaccion.moneda.simbolo}")
            current_row = add_detail_row("Cotización para compra:", f"Gs. {transaccion.cotizacion:,.0f}")
            current_row = add_detail_row("Precio base para compra:", f"Gs. {transaccion.precio_base:,.0f}")
        else:
            current_row = add_detail_row("Monto ingresado para venta:", f"{monto_original_formateado} {transaccion.moneda.simbolo}")
            current_row = add_detail_row("Monto a vender:", f"{monto_formateado} {transaccion.moneda.simbolo}")
            current_row = add_detail_row("Cotización para venta:", f"Gs. {transaccion.cotizacion:,.0f}")
            current_row = add_detail_row("Precio base para venta:", f"Gs. {transaccion.precio_base:,.0f}")
            
            if transaccion.medio_pago == 'Efectivo':
                current_row = add_detail_row("Redondeo por medio pago Efectivo:", f"{redondeo_monto_formateado} {transaccion.moneda.simbolo}")
        
        # Beneficios y recargos
        if transaccion.beneficio_segmento and transaccion.beneficio_segmento > 0:
            current_row = add_detail_row("Beneficio por segmento:", f"Gs. {transaccion.beneficio_segmento:,.0f}")
        
        
        current_row = add_detail_row(f"Recargo por medio de pago ({transaccion.porc_recargo_pago}):", f"Gs. {transaccion.recargo_pago:,.0f}")
        
        
        if transaccion.tipo == 'venta':
            current_row = add_detail_row(f"Recargo por medio de cobro ({transaccion.porc_recargo_cobro}):", f"Gs. {transaccion.recargo_cobro:,.0f}")
        
        # Redondeo efectivo para precio final si aplica
        if ((transaccion.tipo == 'compra' and transaccion.medio_pago == 'Efectivo') or 
            (transaccion.tipo == 'venta' and transaccion.medio_cobro == 'Efectivo')):
            current_row = add_detail_row("Redondeo efectivo precio final:", f"Gs. {transaccion.redondeo_efectivo_precio_final:,.0f}")
        
        # Montos finales
        if transaccion.tipo == 'compra':
            estado_texto = 'pagado' if transaccion.estado in ['Completa', 'Confirmada'] else 'a pagar'
            recibido_texto = 'recibido' if transaccion.estado == 'Completa' else 'a recibir'
            
            current_row = add_detail_row(f"Monto {estado_texto}:", f"Gs. {transaccion.precio_final:,.0f}")
            current_row = add_detail_row(f"Monto {recibido_texto}:", f"{monto_formateado} {transaccion.moneda.simbolo}")
        else:
            estado_texto = 'pagado' if transaccion.estado in ['Completa', 'Confirmada'] else 'a pagar'
            recibido_texto = 'recibido' if transaccion.estado == 'Completa' else 'a recibir'
            
            current_row = add_detail_row(f"Monto {estado_texto}:", f"{monto_formateado} {transaccion.moneda.simbolo}")
            current_row = add_detail_row(f"Monto {recibido_texto}:", f"Gs. {transaccion.precio_final:,.0f}")
        
        # Mostrar monto pagado si es parcial
        if ((transaccion.tipo == 'compra' and transaccion.pagado < transaccion.precio_final and transaccion.estado != 'Completa') or
            (transaccion.tipo == 'venta' and transaccion.pagado < transaccion.monto and transaccion.estado != 'Completa')):
            if transaccion.tipo == 'compra':
                current_row = add_detail_row("Monto pagado:", f"Gs. {transaccion.pagado:,.0f}")
            else:
                current_row = add_detail_row("Monto pagado:", f"{transaccion.pagado:.{transaccion.moneda.decimales}f} {transaccion.moneda.simbolo}")
        
        # Token si existe y está pendiente/confirmada
        if transaccion.token and transaccion.estado in ['Pendiente', 'Confirmada']:
            current_row = add_detail_row("Código de transacción:", transaccion.token)

        
        # Fila de separación entre transacciones
        current_row += 1
    
    # Ajustar ancho de columnas para que coincidan proporcionalmente con detalles
    # Proporcionalmente similar a PDF: más ancho para fecha, iguales para el resto
    column_widths = [25, 18, 20, 15, 15]  # Fecha, Hora, Operación, Moneda, Estado
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Ajustar altura de filas para mejor legibilidad
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 20
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Preparar la respuesta
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"historial_transacciones_{cliente.nombre.replace(' ', '_')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response